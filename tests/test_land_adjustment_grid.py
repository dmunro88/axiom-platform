"""
tests/test_land_adjustment_grid.py — Axiom Platform

Regression coverage for the land_adjustment_grid tab (Phase 6 Adjustment
Grid, added 2026-07-10) and narrative_generator._read_land_adj /
_prompt_land_adjustment, which were rewritten to read it.

Context: the previous `_read_land_adj` read the *wrong* row range (5-14)
from the old 3-section `land` tab -- the real "ADJUSTMENT INPUTS" data
actually lived at rows 31-40. Row 14 held the second section's own header
text ("Sale No.", "Sale Date", ...), which defeated the placeholder-skip
heuristic and crashed with `ValueError: invalid literal for int() with
base 10: 'Sale No.'` whenever realistic data was entered. That whole 3
section `land` tab has since been replaced by the single consolidated
land_adjustment_grid tab (mirroring sca_adjustment_grid's structure), so
this file locks in correct behavior against the new tab rather than
patching the old row-range bug in place.
"""

import sys
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl

import narrative_generator as ng


def _make_workbook(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "Intake"
    ws = wb.create_sheet("land_adjustment_grid")
    for col, text in {
        "A": "Comp", "B": "Location", "I": "Time Adj %",
        "K": "Location Adj %", "L": "Topography Adj %",
        "M": "Surrounding Land Uses Adj %",
    }.items():
        ws[f"{col}6"] = text
    workbook_path = tmp_path / "workbook.xlsx"
    wb.save(workbook_path)
    wb.close()
    return workbook_path


class ReadLandAdjTests(unittest.TestCase):
    def test_reads_populated_rows_with_correct_categories(self):
        with _tmp_dir() as tmp_path:
            workbook_path = _make_workbook(tmp_path)
            wb = openpyxl.load_workbook(workbook_path)
            ws = wb["land_adjustment_grid"]
            ws["B7"] = "123 Fictional Rd"
            ws["I7"] = 0.02
            ws["K7"] = 0.05
            ws["L7"] = -0.03
            ws["M7"] = 0.0
            ws["B8"] = "456 Sample Ave"
            ws["K8"] = -0.10
            wb.save(workbook_path)
            wb.close()

            comps = ng._read_land_adj(workbook_path)

            self.assertEqual(2, len(comps))
            self.assertEqual(1, comps[0]["comp_no"])
            self.assertEqual("123 Fictional Rd", comps[0]["location"])
            self.assertEqual(0.02, comps[0]["time_pct"])
            self.assertEqual(0.05, comps[0]["location_pct"])
            self.assertEqual(-0.03, comps[0]["topo_pct"])
            self.assertEqual(2, comps[1]["comp_no"])
            self.assertEqual(-0.10, comps[1]["location_pct"])
            self.assertEqual(0.0, comps[1]["topo_pct"])

    def test_skips_fully_blank_rows(self):
        with _tmp_dir() as tmp_path:
            workbook_path = _make_workbook(tmp_path)
            # No data entered at all -- every comp row blank.
            comps = ng._read_land_adj(workbook_path)
            self.assertEqual([], comps)

    def test_row_with_only_adjustment_and_no_location_is_still_read(self):
        """A comp with an adjustment entered but no location text yet
        (mid-data-entry) must still surface -- not silently drop real
        adjustment data just because location hasn't been typed in."""
        with _tmp_dir() as tmp_path:
            workbook_path = _make_workbook(tmp_path)
            wb = openpyxl.load_workbook(workbook_path)
            ws = wb["land_adjustment_grid"]
            ws["K7"] = 0.02  # adjustment present, location blank
            wb.save(workbook_path)
            wb.close()

            comps = ng._read_land_adj(workbook_path)
            self.assertEqual(1, len(comps))
            self.assertEqual(0.02, comps[0]["location_pct"])

    def test_missing_tab_returns_empty_list_not_error(self):
        with _tmp_dir() as tmp_path:
            wb = openpyxl.Workbook()
            wb.active.title = "Intake"
            workbook_path = tmp_path / "workbook.xlsx"
            wb.save(workbook_path)
            wb.close()

            comps = ng._read_land_adj(workbook_path)
            self.assertEqual([], comps)

    def test_percentage_point_values_normalized_to_fractions(self):
        """A comp entered as 5.0 (percentage-point convention) rather than
        0.05 (fraction convention) must still normalize correctly, matching
        the same tolerance the old function had."""
        with _tmp_dir() as tmp_path:
            workbook_path = _make_workbook(tmp_path)
            wb = openpyxl.load_workbook(workbook_path)
            ws = wb["land_adjustment_grid"]
            ws["B7"] = "123 Fictional Rd"
            ws["K7"] = 5.0  # entered as a percentage-point number, not 0.05
            wb.save(workbook_path)
            wb.close()

            comps = ng._read_land_adj(workbook_path)
            self.assertEqual(0.05, comps[0]["location_pct"])


class PromptLandAdjustmentTests(unittest.TestCase):
    def test_uses_new_category_names_not_dilmore_or_other(self):
        comps = [{
            "comp_no": 1, "location": "123 Fictional Rd",
            "time_pct": 0.02, "location_pct": 0.05,
            "topo_pct": -0.03, "surrounding_pct": 0.0,
        }]
        prompt, _ = ng._prompt_land_adjustment(comps, {"SUBJECT_ADDRESS": "Subject"})

        self.assertIn("Topography", prompt)
        self.assertIn("Surrounding Land Uses", prompt)
        self.assertNotIn("Dilmore", prompt)
        self.assertNotIn("Size (Dilmore)", prompt)

    def test_zero_adjustments_reported_as_no_adjustment(self):
        comps = [{
            "comp_no": 1, "location": "123 Fictional Rd",
            "time_pct": 0.0, "location_pct": 0.0,
            "topo_pct": 0.0, "surrounding_pct": 0.0,
        }]
        prompt, _ = ng._prompt_land_adjustment(comps, {})
        self.assertIn("Market Conditions: no adjustment", prompt)
        self.assertIn("Location: no adjustment", prompt)


def _tmp_dir():
    import tempfile
    class _Ctx:
        def __enter__(self):
            self._td = tempfile.TemporaryDirectory()
            return Path(self._td.name)
        def __exit__(self, *a):
            self._td.cleanup()
    return _Ctx()


if __name__ == "__main__":
    unittest.main()
