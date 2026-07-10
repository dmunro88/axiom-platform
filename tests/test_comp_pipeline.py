import json
import shutil
import sqlite3
import tempfile
import unittest
from pathlib import Path

import openpyxl

from comp_builder import COMP_COLUMNS, load_comp_data
from comp_library import (
    export_sale_comps_csv,
    export_sale_comps_to_workbook,
)
from comparable_contract import (
    CONTRACT_ID,
    REVIEW_STATUSES,
    SCHEMA_VERSION,
    canonicalize_extraction_result,
    canonicalize_record,
    comparable_identity,
    confirm_extraction_result,
)
from db import (
    backfill_legacy_identities,
    get_conn,
    init_db,
    search_lease_comps,
    search_sale_comps,
)
from ingest import commit_confirmed, commit_extraction_result, run_extraction


def _build_historical_comp_workbook(path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sales"
    sheet.append([
        "Address",
        "City",
        "Property Type",
        "Sale Price",
        "Sale Date",
        "GBA",
        "Price/SF",
        "Cap Rate",
        "Verification",
    ])
    sheet.append([
        "100 Fictional Archive Way",
        "Demo City",
        "Office",
        "$1,000,000",
        "01/15/2025",
        "10,000 SF",
        "$100.00",
        "8.5%",
        "Fictional QA source",
    ])
    lease_sheet = workbook.create_sheet("Lease Comps")
    lease_sheet.append([
        "Address",
        "City",
        "Property Type",
        "Tenant",
        "Lease Date",
        "SF Leased",
        "Rental Rate",
        "Rent Structure",
    ])
    lease_sheet.append([
        "300 Fictional Lease Lane",
        "Demo City",
        "Office",
        "Example Tenant LLC",
        "03/01/2025",
        "2,500 SF",
        "$21.50",
        "Modified Gross",
    ])
    sheet.append([
        "200 Fictional Archive Way",
        "Demo City",
        "Office",
        "$1,000,000",
        "02/20/2025",
        "8,000 SF",
        "$125.00",
        "9.0%",
        "Fictional QA source",
    ])
    # Exact duplicate of row one: should collapse by transaction identity.
    sheet.append([
        "100 Fictional Archive Way",
        "Demo City",
        "Office",
        "$1,000,000",
        "01/15/2025",
        "10,000 SF",
        "$100.00",
        "8.5%",
        "Fictional QA source",
    ])
    workbook.save(path)
    workbook.close()


def _build_comp_export_workbook(path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "comp_data"
    max_column = max(
        ord(column) - ord("A") + 1
        if len(column) == 1
        else (ord(column[0]) - ord("A") + 1) * 26
        + ord(column[1])
        - ord("A")
        + 1
        for column in COMP_COLUMNS.values()
    )
    for key, column in COMP_COLUMNS.items():
        from comp_builder import col_letter_to_idx

        sheet.cell(1, col_letter_to_idx(column), key)
        sheet.cell(2, col_letter_to_idx(column), f"(e.g. {key})")
    self_check = sheet.max_column
    if self_check != max_column:
        raise AssertionError((self_check, max_column))
    workbook.save(path)
    workbook.close()


class ComparablePipelineTests(unittest.TestCase):
    def test_committed_comparable_schema_matches_runtime_contract(self):
        schema_path = (
            Path(__file__).resolve().parents[1]
            / "schemas"
            / "comparable_record.v1.json"
        )
        schema = json.loads(schema_path.read_text(encoding="utf-8"))
        self.assertEqual(CONTRACT_ID, schema["contract_id"])
        self.assertEqual(SCHEMA_VERSION, schema["schema_version"])
        self.assertEqual(
            REVIEW_STATUSES,
            frozenset(schema["review_statuses"]),
        )

    def test_legacy_database_adds_identity_and_review_columns(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "legacy.db"
            connection = sqlite3.connect(db_path)
            connection.executescript(
                """
                CREATE TABLE source_documents (
                    doc_id INTEGER PRIMARY KEY,
                    filename TEXT NOT NULL,
                    filepath TEXT,
                    reviewed INTEGER DEFAULT 0
                );
                CREATE TABLE comps (
                    comp_id INTEGER PRIMARY KEY,
                    property_id INTEGER,
                    source_doc_id INTEGER,
                    sale_date TEXT,
                    reviewed INTEGER DEFAULT 0
                );
                CREATE TABLE lease_comps (
                    lease_comp_id INTEGER PRIMARY KEY,
                    property_id INTEGER,
                    source_doc_id INTEGER,
                    lease_date TEXT,
                    reviewed INTEGER DEFAULT 0
                );
                INSERT INTO comps (comp_id, reviewed) VALUES (1, 1);
                INSERT INTO lease_comps (lease_comp_id, reviewed) VALUES (1, 1);
                """
            )
            connection.commit()
            connection.close()

            init_db(db_path, quiet=True)
            connection = sqlite3.connect(db_path)
            try:
                for table in ("source_documents", "comps", "lease_comps"):
                    columns = {
                        row[1]
                        for row in connection.execute(
                            f"PRAGMA table_info({table})"
                        ).fetchall()
                    }
                    self.assertIn(
                        "content_sha256" if table == "source_documents"
                        else "identity_key",
                        columns,
                    )
                self.assertEqual(
                    "confirmed",
                    connection.execute(
                        "SELECT review_status FROM comps WHERE comp_id = 1"
                    ).fetchone()[0],
                )
                self.assertEqual(
                    "confirmed",
                    connection.execute(
                        """
                        SELECT review_status FROM lease_comps
                        WHERE lease_comp_id = 1
                        """
                    ).fetchone()[0],
                )
            finally:
                connection.close()

    def test_legacy_comp_rows_backfill_identity_key_matching_fresh_import(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "legacy-real-schema.db"
            init_db(db_path, quiet=True)
            connection = get_conn(db_path)
            try:
                cursor = connection.execute(
                    "INSERT INTO properties (address_street, address_city) "
                    "VALUES (?, ?)",
                    ("100 Fictional Legacy Way", "Demo City"),
                )
                property_id = cursor.lastrowid
                # Simulate a row committed before identity_key existed: an
                # ordinary insert with identity_key left NULL, the same as a
                # real pre-upgrade comps row would be.
                connection.execute(
                    "INSERT INTO comps "
                    "(property_id, sale_price, sale_date, reviewed) "
                    "VALUES (?, ?, ?, 1)",
                    (property_id, 1_000_000, "2025-01-15"),
                )
                connection.commit()
            finally:
                connection.close()

            connection = get_conn(db_path)
            try:
                before = connection.execute(
                    "SELECT identity_key FROM comps"
                ).fetchone()["identity_key"]
                self.assertIsNone(before)

                counts = backfill_legacy_identities(connection)
                connection.commit()
                self.assertEqual(1, counts["comps"])

                backfilled = connection.execute(
                    "SELECT identity_key FROM comps"
                ).fetchone()["identity_key"]
                self.assertIsNotNone(backfilled)

                # A fresh import of the identical transaction must compute
                # the exact same identity_key, so `comparable_id_by_identity`
                # correctly recognizes it as already-ingested instead of
                # inserting a duplicate.
                fresh_identity = comparable_identity("sale", {
                    "address_street": "100 Fictional Legacy Way",
                    "address_city": "Demo City",
                    "sale_date": "2025-01-15",
                    "sale_price": 1_000_000,
                })
                self.assertEqual(fresh_identity, backfilled)

                # Calling it again must not change an already-backfilled row.
                second_counts = backfill_legacy_identities(connection)
                self.assertEqual(0, second_counts["comps"])
            finally:
                connection.close()

    def test_identity_does_not_collapse_distinct_sales_at_same_price(self):
        first = {
            "address_street": "100 Fictional Way",
            "address_city": "Demo City",
            "sale_date": "2025-01-01",
            "sale_price": 1_000_000,
        }
        second = dict(first, address_street="200 Fictional Way")
        duplicate = dict(first, address_street=" 100 FICTIONAL WAY ")
        self.assertNotEqual(
            comparable_identity("sale", first),
            comparable_identity("sale", second),
        )
        self.assertEqual(
            comparable_identity("sale", first),
            comparable_identity("sale", duplicate),
        )

    def test_placeholder_lease_expiration_canonicalizes_to_blank(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "lease-source.xlsx"
            source.write_text("fictional source", encoding="utf-8")
            record = canonicalize_record(
                {
                    "data": {
                        "address_street": "300 Fictional Lease Lane",
                        "base_rent_psf": "$21.50",
                        "lease_expiration": "N/A",
                    },
                    "confidence": {
                        "address_street": "high",
                        "base_rent_psf": "high",
                        "lease_expiration": "high",
                    },
                    "source": str(source),
                },
                "lease",
                source_path=source,
            )

        self.assertIsNone(record["data"]["lease_expiration"])
        self.assertEqual([], record["validation"]["errors"])

    def test_fictional_extract_review_commit_search_and_export(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            projects = root / "historical"
            assignment = (
                projects
                / "25C001 - Office - 999 Fictional Subject Road, Demo City"
            )
            assignment.mkdir(parents=True)
            source_path = assignment / "Market Chart.xlsx"
            _build_historical_comp_workbook(source_path)

            staged_dir = root / "ingest" / "staged"
            staged_paths = run_extraction(projects, staged_dir=staged_dir)
            self.assertEqual(1, len(staged_paths))
            staged = json.loads(staged_paths[0].read_text(encoding="utf-8"))
            self.assertEqual(
                "axiom.comparable.extraction_batch",
                staged["contract_id"],
            )
            self.assertEqual(SCHEMA_VERSION, staged["schema_version"])
            self.assertEqual(2, len(staged["comps"]))
            self.assertEqual(1, len(staged["lease_comps"]))
            self.assertEqual(
                [0.085, 0.09],
                [record["data"]["cap_rate"] for record in staged["comps"]],
            )
            for record in staged["comps"]:
                self.assertEqual(CONTRACT_ID, record["contract_id"])
                self.assertTrue(record["identity_key"])
                self.assertEqual(64, len(record["provenance"]["source_sha256"]))
                self.assertEqual("unreviewed", record["review"]["status"])

            staged["comps"][0]["data"][
                "verification_source"
            ] = "Reviewed fictional QA source"
            staged["comps"][0]["review_edits"] = [
                {
                    "field": "verification_source",
                    "before": "Fictional QA source",
                    "after": "Reviewed fictional QA source",
                }
            ]
            confirmed = confirm_extraction_result(
                staged,
                reviewer="fictional-test-reviewer",
                reviewed_at="2026-07-01T12:00:00+00:00",
            )
            self.assertEqual(
                "verification_source",
                confirmed["comps"][0]["review"]["edits"][0]["field"],
            )
            confirmed_dir = root / "ingest" / "confirmed"
            confirmed_dir.mkdir(parents=True)
            confirmed_path = confirmed_dir / "batch.json"
            confirmed_path.write_text(
                json.dumps(confirmed, indent=2),
                encoding="utf-8",
            )

            db_path = root / "axiom-test.db"
            summary = commit_confirmed(
                confirmed_dir=confirmed_dir,
                db_path=db_path,
            )
            self.assertEqual(2, summary["sale_comps"])
            self.assertEqual(1, summary["lease_comps"])
            self.assertTrue(confirmed_path.with_suffix(".committed").exists())

            records = search_sale_comps(
                db_path,
                city="Demo City",
                property_type="Office",
            )
            self.assertEqual(2, len(records))
            self.assertEqual(
                {
                    "100 Fictional Archive Way",
                    "200 Fictional Archive Way",
                },
                {record["address_street"] for record in records},
            )
            self.assertTrue(
                all(record["review_status"] == "confirmed" for record in records)
            )
            lease_records = search_lease_comps(
                db_path,
                city="Demo City",
                property_type="Office",
            )
            self.assertEqual(1, len(lease_records))
            self.assertEqual(
                "Example Tenant LLC",
                lease_records[0]["tenant_name"],
            )
            self.assertEqual(21.5, lease_records[0]["base_rent_psf"])

            # Recommit identical source content from a different path.
            moved_source = root / "moved" / "Renamed Archive.xlsx"
            moved_source.parent.mkdir()
            shutil.copy2(source_path, moved_source)
            moved_batch = json.loads(json.dumps(staged))
            moved_batch["sources"] = [str(moved_source)]
            for record in moved_batch["comps"]:
                record["source"] = str(moved_source)
                record["provenance"] = {}
            moved_confirmed = confirm_extraction_result(
                moved_batch,
                reviewer="fictional-test-reviewer",
                reviewed_at="2026-07-02T12:00:00+00:00",
            )
            second_path = confirmed_dir / "same-content-new-name.json"
            second_path.write_text(
                json.dumps(moved_confirmed, indent=2),
                encoding="utf-8",
            )
            second_summary = commit_confirmed(
                confirmed_dir=confirmed_dir,
                db_path=db_path,
            )
            self.assertEqual(0, second_summary["sale_comps"])

            connection = sqlite3.connect(db_path)
            try:
                self.assertEqual(
                    1,
                    connection.execute(
                        "SELECT COUNT(*) FROM source_documents"
                    ).fetchone()[0],
                )
                self.assertEqual(
                    2,
                    connection.execute("SELECT COUNT(*) FROM comps").fetchone()[0],
                )
                self.assertEqual(
                    1,
                    connection.execute(
                        "SELECT COUNT(*) FROM lease_comps"
                    ).fetchone()[0],
                )
                self.assertEqual(
                    3,
                    connection.execute(
                        "SELECT COUNT(*) FROM properties"
                    ).fetchone()[0],
                )
            finally:
                connection.close()

            workbook_template = root / "workbook-template.xlsx"
            workbook_output = root / "workbook-with-comps.xlsx"
            _build_comp_export_workbook(workbook_template)
            export_sale_comps_to_workbook(
                records,
                workbook_template,
                workbook_output,
            )
            report_comps = load_comp_data(workbook_output)
            self.assertEqual(2, len(report_comps))
            self.assertEqual("Sale No. 1", report_comps[0]["COMP_NO"])
            self.assertIn(
                report_comps[0]["COMP_ADDRESS_LINE1"],
                {
                    "100 Fictional Archive Way",
                    "200 Fictional Archive Way",
                },
            )

            csv_path = export_sale_comps_csv(records, root / "comps.csv")
            csv_text = csv_path.read_text(encoding="utf-8-sig")
            self.assertIn("Fictional Archive Way", csv_text)

    def test_unreviewed_batch_rolls_back_without_partial_rows(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "fictional.xlsx"
            _build_historical_comp_workbook(source)
            batch = canonicalize_extraction_result({
                "folder_name": "Fictional Unreviewed",
                "folder_meta": {},
                "sources": [str(source)],
                "narrative": {},
                "income_data": {},
                "comps": [
                    {
                        "data": {
                            "address_street": "1 Unreviewed Way",
                            "sale_price": 100000,
                        },
                        "confidence": {
                            "address_street": "high",
                            "sale_price": "high",
                        },
                        "source": str(source),
                    }
                ],
                "lease_comps": [],
            })
            db_path = root / "rollback.db"
            init_db(db_path)
            connection = get_conn(db_path)
            try:
                with self.assertRaisesRegex(ValueError, "not been confirmed"):
                    with connection:
                        commit_extraction_result(batch, connection)
                for table in ("source_documents", "properties", "comps"):
                    self.assertEqual(
                        0,
                        connection.execute(
                            f"SELECT COUNT(*) FROM {table}"
                        ).fetchone()[0],
                    )
            finally:
                connection.close()

    def test_unconfirmed_comp_in_confirmed_batch_raises_instead_of_silent_skip(
        self,
    ):
        """A hand-edited or otherwise non-standard confirmed file could carry
        an outer batch marked "confirmed" while one comp record inside it
        still isn't. commit_extraction_result() used to silently `continue`
        past that record -- so a batch where every comp was individually
        unconfirmed could still get marked ".committed" by commit_confirmed()
        while contributing zero database rows, with no warning. This now
        raises, matching every other harvest record type (rent_roll/expense/
        observation/artifact), instead of silently skipping."""
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "fictional.xlsx"
            _build_historical_comp_workbook(source)
            batch = canonicalize_extraction_result({
                "folder_name": "Fictional Unconfirmed Comp",
                "folder_meta": {},
                "sources": [str(source)],
                "narrative": {},
                "income_data": {},
                "comps": [
                    {
                        "data": {
                            "address_street": "1 Unconfirmed Comp Way",
                            "sale_price": 100000,
                        },
                        "confidence": {
                            "address_street": "high",
                            "sale_price": "high",
                        },
                        "source": str(source),
                    }
                ],
                "lease_comps": [],
            })
            # Mark the outer batch confirmed but leave the individual comp's
            # own review status untouched (still "unreviewed"), mimicking a
            # corrupted/hand-edited confirmed file rather than the normal
            # review_staged()/confirm_extraction_result() flow, which would
            # have marked both consistently.
            batch["review"] = {"status": "confirmed"}
            db_path = root / "unconfirmed_comp.db"
            init_db(db_path)
            connection = get_conn(db_path)
            try:
                with self.assertRaisesRegex(
                    ValueError,
                    "has not been confirmed",
                ):
                    with connection:
                        commit_extraction_result(batch, connection)
                for table in ("source_documents", "properties", "comps"):
                    self.assertEqual(
                        0,
                        connection.execute(
                            f"SELECT COUNT(*) FROM {table}"
                        ).fetchone()[0],
                    )
            finally:
                connection.close()

    def test_source_change_after_extraction_requires_reextract(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "fictional.xlsx"
            _build_historical_comp_workbook(source)
            batch = confirm_extraction_result(
                canonicalize_extraction_result({
                    "folder_name": "Fictional Changed Source",
                    "folder_meta": {},
                    "sources": [str(source)],
                    "narrative": {},
                    "income_data": {},
                    "comps": [
                        {
                            "data": {
                                "address_street": "1 Changed Source Way",
                                "sale_price": 100000,
                            },
                            "confidence": {
                                "address_street": "high",
                                "sale_price": "high",
                            },
                            "source": str(source),
                        }
                    ],
                    "lease_comps": [],
                }),
                reviewer="fictional-reviewer",
                reviewed_at="2026-07-01T12:00:00+00:00",
            )
            with open(source, "ab") as source_file:
                source_file.write(b"changed-after-extraction")

            db_path = root / "changed.db"
            init_db(db_path)
            connection = get_conn(db_path)
            try:
                with self.assertRaisesRegex(
                    ValueError,
                    "Source changed after extraction",
                ):
                    with connection:
                        commit_extraction_result(batch, connection)
                self.assertEqual(
                    0,
                    connection.execute(
                        "SELECT COUNT(*) FROM source_documents"
                    ).fetchone()[0],
                )
            finally:
                connection.close()


if __name__ == "__main__":
    unittest.main()
