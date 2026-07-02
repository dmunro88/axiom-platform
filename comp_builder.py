"""
comp_builder.py — Axiom Commercial Appraisal
Reads comp data from the workbook's comp_data sheet,
fills the comp_block_template for each comp, and injects
the resulting pages into the appraisal report at [[COMP_SHEETS_BLOCK]].

Called automatically by axiom.py during the deliver stage.
Can also be run standalone:
    python comp_builder.py <report_path> <workbook_path> <template_dir>
"""

import copy, sys
from pathlib import Path

from docx import Document
from docx.oxml.ns import qn
import openpyxl


# ── Column map for comp_data sheet ───────────────────────────────────────────
# Each key = [[COMP_*]] placeholder (without brackets)
# Each value = column letter in the comp_data sheet
COMP_COLUMNS = {
    "COMP_NO":              "A",   # e.g. "Sale No. 1"
    "COMP_SUBMARKET":       "B",   # e.g. "Hoover"
    "COMP_ADDRESS_LINE1":   "C",   # e.g. "437-439 Industrial Lane"
    "COMP_ADDRESS_LINE2":   "D",   # e.g. "Homewood, Alabama 35216"
    "COMP_PROPERTY_TYPE":   "E",   # e.g. "Multi-Tenant Office"
    "COMP_SALE_PRICE":      "F",   # e.g. "$892,000"
    "COMP_SALE_DATE":       "G",   # e.g. "12/26/2025"
    "COMP_GBA_SF":          "H",   # e.g. "9,780 SF"
    "COMP_PRICE_SF":        "I",   # e.g. "$91.26"
    "COMP_CAP_RATE":        "J",   # e.g. "8.10%"
    "COMP_YEAR_BUILT":      "K",   # e.g. "1988"
    "COMP_SITE_AREA":       "L",   # e.g. "± 110,002 SF"
    "COMP_STORIES":         "M",   # e.g. "One"
    "COMP_CONSTRUCTION":    "N",   # e.g. "Metal"
    "COMP_CONDITION":       "O",   # e.g. "Average"
    "COMP_ZONING":          "P",   # e.g. "Office/Commercial"
    "COMP_TOPOGRAPHY":      "Q",   # e.g. "Level"
    "COMP_SHAPE":           "R",   # e.g. "Irregular"
    "COMP_FLOOD_ZONE":      "S",   # e.g. "Zone X"
    "COMP_ACCESS":          "T",   # e.g. "Average"
    "COMP_VISIBILITY":      "U",   # e.g. "Average"
    "COMP_UTILITIES":       "V",   # e.g. "All Public"
    "COMP_GRANTOR":         "W",   # e.g. "AMC Real Estate Holdings LLC"
    "COMP_GRANTEE":         "X",   # e.g. "5690 Pine Lane Cir, LLC"
    "COMP_DEED_REF":        "Y",   # e.g. "2026002021"
    "COMP_VERIFICATION":    "Z",   # e.g. "CoStar / Public Records"
    "COMP_NOI":             "AA",  # e.g. "$72,293"
    "COMP_NOI_SF":          "AB",  # e.g. "$7.39"
    "COMP_ANALYSIS":        "AC",  # main analysis paragraph
    "COMP_RELEVANCE":       "AD",  # relevance sentence
    "COMP_POPULATION":      "AE",  # e.g. "132,176"
    "COMP_HH_INCOME":       "AF",  # e.g. "$94,659"
    "COMP_LIST_PRICE":      "AG",  # e.g. "$1,000,000"
    "COMP_PCT_DISCOUNT":    "AH",  # e.g. "11%"
    "COMP_TIME_ON_MARKET":  "AI",  # e.g. "52 days"
}

COL_TO_KEY = {v: k for k, v in COMP_COLUMNS.items()}
# Build col-letter → col-index map (A=1, B=2, ..., Z=26, AA=27, AB=28, ...)
def col_letter_to_idx(col):
    idx = 0
    for ch in col.upper():
        idx = idx * 26 + (ord(ch) - ord('A') + 1)
    return idx

COL_INDICES = {col: col_letter_to_idx(col) for col in COMP_COLUMNS.values()}


# ── Load comp data from workbook ──────────────────────────────────────────────

def load_comp_data(workbook_path):
    """
    Read comp_data sheet from workbook.
    Returns list of dicts, one per comp row (rows with a COMP_NO value).
    """
    wb = openpyxl.load_workbook(str(workbook_path), data_only=True)
    if "comp_data" not in wb.sheetnames:
        print("  Warning: no comp_data sheet found in workbook — skipping comp pages.")
        return []

    ws = wb["comp_data"]
    comps = []

    # Row 1 is the header row; data starts at row 2
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Check if comp number column (A=index 0) has a real value (not example/header)
        comp_no_val = row[0]
        if not comp_no_val:
            continue
        # Skip template example/header rows
        if '(e.g.' in str(comp_no_val) or 'COMP_NO' in str(comp_no_val):
            continue
        # Require at least one core data field (skip blank placeholder rows)
        sale_price = row[col_letter_to_idx('F') - 1] if col_letter_to_idx('F') - 1 < len(row) else None
        address    = row[col_letter_to_idx('C') - 1] if col_letter_to_idx('C') - 1 < len(row) else None
        if not sale_price and not address:
            continue

        comp = {}
        for col_letter, key in COL_TO_KEY.items():
            col_idx = COL_INDICES[col_letter] - 1  # 0-based
            if col_idx < len(row) and row[col_idx] is not None:
                comp[key] = str(row[col_idx]).strip()
            else:
                comp[key] = ""

        comps.append(comp)

    return comps


# ── Placeholder fill in a cloned block ───────────────────────────────────────

def _fill_text_node(t_el, comp):
    """Replace [[COMP_*]] in a w:t element with comp data."""
    if not t_el.text:
        return
    text = t_el.text
    changed = False
    for key, val in comp.items():
        placeholder = f"[[{key}]]"
        if placeholder in text:
            text = text.replace(placeholder, val)
            changed = True
    if changed:
        t_el.text = text
        if text and (text[0] == ' ' or text[-1] == ' '):
            t_el.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')


def fill_comp_block(block_elements, comp):
    """
    Fill all [[COMP_*]] placeholders in a list of XML elements (cloned).
    Returns the filled elements.
    """
    filled = [copy.deepcopy(el) for el in block_elements]
    for el in filled:
        for t in el.iter(qn('w:t')):
            _fill_text_node(t, comp)
    return filled


# ── Page break paragraph ──────────────────────────────────────────────────────

def make_page_break_para():
    """Create an empty paragraph with a page break."""
    from lxml import etree
    W = 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'
    p = etree.Element(qn('w:p'))
    r = etree.SubElement(p, qn('w:r'))
    br = etree.SubElement(r, qn('w:br'))
    br.set(qn('w:type'), 'page')
    return p


# ── Inject comp pages into report ─────────────────────────────────────────────

def inject_comp_section(report_path, comp_template_path, workbook_path):
    """
    Main entry point.
    - Finds [[COMP_SHEETS_BLOCK]] in report_path
    - Loads comp data from workbook
    - For each comp, fills the template and inserts at the marker
    - Saves the modified report in place
    Returns the number of comps injected.
    """
    comps = load_comp_data(workbook_path)
    if not comps:
        print("  No comp data found — [[COMP_SHEETS_BLOCK]] left unfilled.")
        return 0

    # Load the comp block template and extract its body elements (excluding sectPr)
    tmpl_doc  = Document(str(comp_template_path))
    tmpl_body = tmpl_doc.element.body
    tmpl_elements = [
        el for el in list(tmpl_body)
        if el.tag != qn('w:sectPr')
    ]
    print(f"  Comp block template: {len(tmpl_elements)} elements per comp")

    # Load the target report
    report_doc  = Document(str(report_path))
    report_body = report_doc.element.body

    # Find the [[COMP_SHEETS_BLOCK]] marker paragraph
    marker_para = None
    for para in report_body.iter(qn('w:p')):
        texts = ''.join(t.text or '' for t in para.iter(qn('w:t')))
        if '[[COMP_SHEETS_BLOCK]]' in texts:
            marker_para = para
            break

    if marker_para is None:
        print("  Warning: [[COMP_SHEETS_BLOCK]] marker not found in report — comp pages not injected.")
        return 0

    # Find marker's position in body
    parent     = marker_para.getparent()
    insert_pos = list(parent).index(marker_para)

    print(f"  Inserting {len(comps)} comp(s) at position {insert_pos}")

    # Insert comp blocks in reverse order so insert_pos stays valid
    # (we insert at the same index each time, pushing earlier inserts down)
    # Actually insert forward: increment insert_pos after each block
    current_pos = insert_pos

    for comp_idx, comp in enumerate(comps):
        # Page break before comp 2, 3, ... (not before comp 1)
        if comp_idx > 0:
            pb = make_page_break_para()
            parent.insert(current_pos, pb)
            current_pos += 1

        # Fill and insert this comp's block elements
        filled = fill_comp_block(tmpl_elements, comp)
        for el in filled:
            parent.insert(current_pos, el)
            current_pos += 1

    # Remove the marker paragraph
    parent.remove(marker_para)

    report_doc.save(str(report_path))
    print(f"  OK: {len(comps)} comp page(s) injected into report.")
    return len(comps)


# ── Standalone entry point ────────────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("Usage: python comp_builder.py <report.docx> <workbook.xlsx> <templates_dir>")
        sys.exit(1)

    report_path        = Path(sys.argv[1])
    workbook_path      = Path(sys.argv[2])
    templates_dir      = Path(sys.argv[3])
    comp_template_path = templates_dir / "comp_block_template.docx"

    if not report_path.exists():
        print(f"Error: report not found: {report_path}")
        sys.exit(1)
    if not workbook_path.exists():
        print(f"Error: workbook not found: {workbook_path}")
        sys.exit(1)
    if not comp_template_path.exists():
        print(f"Error: comp template not found: {comp_template_path}")
        sys.exit(1)

    n = inject_comp_section(report_path, comp_template_path, workbook_path)
    print(f"Done. {n} comp page(s) injected.")
