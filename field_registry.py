"""Versioned field-contract inventory and drift auditing for Axiom."""

import json
import re
import zipfile
from pathlib import Path

import openpyxl
from docx.oxml.ns import qn
from lxml import etree

from presentation_variants import VARIANT_RULES


KEY_PATTERN = re.compile(r"^[A-Z][A-Z0-9_]*$")
PLACEHOLDER_PATTERN = re.compile(r"\[\[([A-Z0-9_]+)\]\]")
BLOCK_SUFFIXES = (
    "_BLOCK",
    "_IMAGE",
    "_NARRATIVE",
    "_OVERVIEW",
    "_MAP",
    "_SECTION",
    "_TABLE",
)

MEDIA_BLOCKS = frozenset({
    "AERIAL_MAP_IMAGE",
    "BUILDING_SKETCH_BLOCK",
    "LEASE_COMP_LOCATION_MAP",
    "LEASE_COMP_PHOTOS_BLOCK",
    "LAND_SALE_LOCATION_MAP",
    "PARCEL_MAP_IMAGE",
    "REGIONAL_MAP_IMAGE",
    "SCA_SALE_LOCATION_MAP",
    "SUBJECT_PHOTOS_BLOCK",
})

TEXT_VALUE_KEYS = frozenset({
    "VALUE_INTEREST",
    "VALUE_INTEREST_LOWER",
    "VALUE_TYPE",
    "VALUE_TYPE_SHORT",
    "VALUE_WORDS",
    "VALUE_WORDS_FORMAL",
})


def _is_key(value):
    return isinstance(value, str) and KEY_PATTERN.fullmatch(value.strip())


def _is_block(key):
    return any(key.endswith(suffix) for suffix in BLOCK_SUFFIXES)


def _output_row_is_produced(row_number, raw_value, formatted_value):
    if raw_value not in (None, ""):
        return True
    if formatted_value in (None, ""):
        return False
    if not (
        isinstance(formatted_value, str)
        and formatted_value.startswith("=")
    ):
        return True

    references = re.findall(
        r"(?:'[^']+'!|[A-Za-z0-9_]+!)?\$?([A-Z]{1,3})\$?(\d+)",
        formatted_value,
    )
    if not references:
        return True
    return any(
        column != "C" or int(referenced_row) != row_number
        for column, referenced_row in references
    )


def _docx_placeholders(docx_path):
    placeholders = set()
    with zipfile.ZipFile(docx_path) as package:
        for name in package.namelist():
            if name.startswith("word/") and name.endswith(".xml"):
                root = etree.fromstring(package.read(name))
                for paragraph in root.iter(qn("w:p")):
                    text = "".join(
                        node.text or ""
                        for node in paragraph.iter(qn("w:t"))
                    )
                    placeholders.update(PLACEHOLDER_PATTERN.findall(text))
    return placeholders


def inventory_workbook(workbook_path):
    """Return Intake and outputs field inventories from an Axiom workbook."""
    workbook = openpyxl.load_workbook(
        workbook_path,
        read_only=True,
        data_only=False,
    )
    intake = {}
    outputs = set()
    produced_outputs = set()
    try:
        if "Intake" in workbook.sheetnames:
            sheet = workbook["Intake"]
            for row in sheet.iter_rows(min_row=1, max_col=3, values_only=True):
                key = row[0]
                if _is_key(key):
                    intake[key.strip()] = str(row[2] or "").strip()

        if "outputs" in workbook.sheetnames:
            sheet = workbook["outputs"]
            for row_number, row in enumerate(
                sheet.iter_rows(
                min_row=1,
                min_col=2,
                max_col=4,
                values_only=True,
                ),
                start=1,
            ):
                key = row[0]
                if _is_key(key):
                    normalized_key = key.strip()
                    outputs.add(normalized_key)
                    if _output_row_is_produced(
                        row_number,
                        row[1],
                        row[2],
                    ):
                        produced_outputs.add(normalized_key)
    finally:
        workbook.close()

    return {
        "intake": intake,
        "workbook_output": outputs,
        "workbook_output_produced": produced_outputs,
    }


def inventory_templates(templates_dir, stages):
    """Return placeholder usage by configured workflow stage."""
    templates_dir = Path(templates_dir)
    usage = {}
    for stage_name, stage in stages.items():
        stage_keys = set()
        for document in stage.get("documents", []):
            template_name = document.get("template")
            template_path = templates_dir / template_name if template_name else None
            if template_path and template_path.exists():
                stage_keys.update(_docx_placeholders(template_path))
        usage[stage_name] = stage_keys
    return usage


def _infer_value_kind(key):
    if key in TEXT_VALUE_KEYS:
        return "text"
    if key.endswith("_DATE"):
        return "date_text"
    if key.endswith(("_RATE", "_RATIO", "_PERCENT", "_WEIGHT")):
        return "percentage_text"
    if any(
        token in key
        for token in (
            "PRICE",
            "VALUE",
            "FEE",
            "COST",
            "RENT",
            "TAX",
            "NOI",
            "EGI",
            "PGI",
        )
    ):
        return "currency_text"
    if key.endswith(("_SF", "_COUNT", "_UNITS", "_FLOORS", "_SPACES")):
        return "number_text"
    if key.startswith(("CA_DEVELOPED", "SCA_DEVELOPED", "IA_DEVELOPED")):
        return "yes_no_text"
    return "text"


ADJUSTMENT_GRID_BLOCKS = frozenset({
    "SCA_ADJUSTMENT_GRID_BLOCK",
    "SCA_QUALITATIVE_GRID_BLOCK",
    "LAND_ADJUSTMENT_GRID_BLOCK",
    "LAND_QUALITATIVE_GRID_BLOCK",
})


def _block_handler(key):
    if key in MEDIA_BLOCKS:
        return "media"
    if key == "OWNERSHIP_HISTORY_TABLE":
        return "structured"
    if key == "COMP_SHEETS_BLOCK" or key in ADJUSTMENT_GRID_BLOCKS:
        return "comparables"
    if key.endswith(("_NARRATIVE", "_OVERVIEW")):
        return "narrative"
    return "unregistered"


def build_registry(
    workbook_path,
    templates_dir,
    stages,
    fixture_json_path=None,
    schema_version="1.2.0",
):
    """Build the initial authoritative registry from the verified baseline."""
    workbook_inventory = inventory_workbook(workbook_path)
    template_usage = inventory_templates(templates_dir, stages)

    fixture_keys = set()
    if fixture_json_path and Path(fixture_json_path).exists():
        with open(fixture_json_path, encoding="utf-8") as fixture_file:
            fixture_keys.update(json.load(fixture_file))

    all_keys = (
        set(workbook_inventory["intake"])
        | workbook_inventory["workbook_output"]
        | fixture_keys
        | set().union(*template_usage.values())
        | set(VARIANT_RULES)
    )

    fields = {}
    blocks = {}
    source_keys = (
        set(workbook_inventory["intake"])
        | workbook_inventory["workbook_output"]
        | fixture_keys
        | set(VARIANT_RULES)
    )
    for key in sorted(all_keys):
        used_in = sorted(
            stage for stage, keys in template_usage.items() if key in keys
        )
        if _is_block(key) and key not in source_keys:
            blocks[key] = {
                "handler": _block_handler(key),
                "used_in": used_in,
            }
            continue

        producers = []
        if key in workbook_inventory["intake"]:
            producers.append("intake")
        if key in workbook_inventory["workbook_output_produced"]:
            producers.append("workbook_output")
        if key in fixture_keys:
            producers.append("json_export")

        variant_rule = VARIANT_RULES.get(key)
        if variant_rule:
            source_of_truth = "application"
            producers = ["application"]
        elif key in workbook_inventory["workbook_output_produced"]:
            source_of_truth = "workbook_output"
        elif key in workbook_inventory["intake"]:
            source_of_truth = "intake"
        elif key in fixture_keys:
            source_of_truth = "json_export"
        else:
            source_of_truth = "application"

        field = {
            "value_kind": _infer_value_kind(key),
            "source_of_truth": source_of_truth,
            "producers": producers,
            "used_in": used_in,
        }
        if variant_rule:
            field["derived_from"] = [variant_rule["source"]]
            field["derivation"] = variant_rule["transform"]
        description = workbook_inventory["intake"].get(key)
        if description:
            field["description"] = description
            if description.strip().lower() == "leave blank if none":
                field["required"] = False
        fields[key] = field

    return {
        "contract_id": "axiom.appraisal.fields",
        "schema_version": schema_version,
        "status": "baseline",
        "load_order": ["json_export", "workbook_output"],
        "precedence_rule": "last_nonblank_value_wins",
        "fields": fields,
        "blocks": blocks,
    }


def load_registry(registry_path):
    """Load and minimally validate a field registry."""
    with open(registry_path, encoding="utf-8") as registry_file:
        registry = json.load(registry_file)
    if registry.get("contract_id") != "axiom.appraisal.fields":
        raise ValueError("Unrecognized field registry contract_id.")
    if not registry.get("schema_version"):
        raise ValueError("Field registry has no schema_version.")
    if not isinstance(registry.get("fields"), dict):
        raise ValueError("Field registry has no fields object.")
    if not isinstance(registry.get("blocks"), dict):
        raise ValueError("Field registry has no blocks object.")
    return registry


def registry_version(registry_path):
    """Return the configured schema version, or ``None`` if unavailable."""
    try:
        return load_registry(registry_path)["schema_version"]
    except (OSError, ValueError, json.JSONDecodeError):
        return None


def audit_assignment_contract(
    registry_path,
    workbook_path,
    template_paths,
    variables,
):
    """Compare live assignment inputs and templates with the registry."""
    result = {
        "schema_version": None,
        "errors": [],
        "warnings": [],
    }
    try:
        registry = load_registry(registry_path)
    except Exception as exc:
        result["errors"].append(f"Field registry could not be loaded: {exc}")
        return result

    result["schema_version"] = registry["schema_version"]
    registered_fields = set(registry["fields"])
    registered_blocks = set(registry["blocks"])
    registered_keys = registered_fields | registered_blocks

    unregistered_handlers = sorted(
        key
        for key, definition in registry["blocks"].items()
        if definition.get("handler") == "unregistered"
        and definition.get("used_in")
    )
    if unregistered_handlers:
        result["errors"].append(
            "Registry blocks have no runtime handler: "
            + ", ".join(unregistered_handlers)
            + "."
        )

    template_keys = set()
    for template_path in template_paths:
        template_keys.update(_docx_placeholders(template_path))
    unknown_template = sorted(template_keys - registered_keys)
    if unknown_template:
        result["errors"].append(
            "Template placeholders are absent from field registry: "
            + ", ".join(unknown_template)
            + "."
        )

    # Reverse direction: a registered block that claims to be used somewhere
    # (used_in is non-empty) must actually have its marker present in one of
    # the audited templates. Without this check, a registered block whose
    # marker is missing or typo'd in the template passes contract clean while
    # silently never appearing in the delivered report -- the exact failure
    # mode the historical outputs-tab key mismatch already hit once.
    missing_from_template = sorted(
        key
        for key, definition in registry["blocks"].items()
        if definition.get("used_in") and key not in template_keys
    )
    if missing_from_template:
        result["errors"].append(
            "Registry blocks are missing their marker in the template: "
            + ", ".join(missing_from_template)
            + "."
        )

    workbook_inventory = inventory_workbook(workbook_path)
    workbook_keys = (
        set(workbook_inventory["intake"])
        | workbook_inventory["workbook_output"]
    )
    unknown_workbook = sorted(workbook_keys - registered_fields)
    if unknown_workbook:
        result["errors"].append(
            "Workbook fields are absent from field registry: "
            + ", ".join(unknown_workbook)
            + "."
        )

    unknown_variables = sorted(
        key
        for key in set(variables) - registered_fields
        if _is_key(key)
    )
    if unknown_variables:
        result["warnings"].append(
            "Assignment JSON contains unregistered fields: "
            + ", ".join(unknown_variables)
            + "."
        )

    return result
