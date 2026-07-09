"""Structured extraction for historical rent rolls and operating expenses."""

import datetime
import math
import re
from pathlib import Path

import openpyxl

from comparable_contract import source_metadata


RENT_ROLL_SYNONYMS = {
    "unit_id": [
        "unit",
        "unit no",
        "unit number",
        "unit #",
        "space",
        "space no",
        "space number",
        "space #",
        "lot",
        "lot no",
        "lot number",
        "lot #",
        "site",
        "site no",
        "site number",
        "site #",
        "room",
        "room no",
        "room number",
        "room #",
        "apt",
        "apt no",
        "apartment",
        "apartment no",
    ],
    "suite": ["suite", "suite no", "suite number"],
    "tenant_name": [
        "tenant",
        "tenant name",
        "lessee",
        "occupant",
        "resident",
        "resident name",
        "name",
    ],
    "tenant_first_name": ["first name", "resident first", "tenant first"],
    "tenant_last_name": ["last name", "resident last", "tenant last"],
    "tenant_use": [
        "use",
        "tenant use",
        "business",
        "space use",
        "unit type",
        "site type",
        "lot type",
        "room type",
        "size",
    ],
    "sf_leased": [
        "sf",
        "square feet",
        "rentable sf",
        "leased sf",
        "suite sf",
        "area",
    ],
    "lease_start": [
        "lease start",
        "start date",
        "commencement",
        "commencement date",
        "move in",
        "move-in",
        "move in date",
        "move-in date",
    ],
    "lease_end": [
        "lease end",
        "end date",
        "expiration",
        "expiration date",
        "lease expiration",
        "lease exp",
        "exp date",
    ],
    "monthly_rent": [
        "monthly rent",
        "rent monthly",
        "rent/mo",
        "base rent/mo",
        "rent",
        "lot rent",
        "site rent",
        "room rent",
        "current rent",
        "rate",
        "payment",
    ],
    "annual_rent": ["annual rent", "rent annual", "base rent annual"],
    "rent_psf": [
        "rent psf",
        "rent/sf",
        "annual rent/sf",
        "base rent psf",
        "rate psf",
    ],
    "reimbursement_structure": [
        "rent structure",
        "lease type",
        "reimbursement",
        "expense structure",
        "lease basis",
    ],
    "occupancy_status": ["status", "occupancy", "occupancy status"],
    "discounts": ["discount", "discounts", "concession", "concessions"],
    "notes": ["notes", "note", "comments", "memo"],
    "balance": ["balance", "amount due", "arrears"],
}

EXPENSE_SYNONYMS = {
    "category": [
        "category",
        "expense",
        "expense category",
        "account",
        "line item",
        "description",
    ],
    "period_year": ["year", "period year", "fiscal year"],
    "period_type": ["period type", "type", "basis", "scenario"],
    "amount": ["amount", "expense amount", "annual amount", "total"],
    "amount_per_sf": ["per sf", "amount/sf", "expense/sf", "$/sf"],
    "notes": ["notes", "comments", "source notes"],
}

RENT_NUMBERS = {
    "sf_leased",
    "monthly_rent",
    "annual_rent",
    "rent_psf",
    "discounts",
    "balance",
}
RENT_DATES = {"lease_start", "lease_end", "as_of_date"}
EXPENSE_NUMBERS = {"period_year", "amount", "amount_per_sf"}
TOTAL_PREFIXES = (
    "total",
    "subtotal",
    "net operating income",
    "effective gross",
)
PER_SF_MARKERS = (
    "$/sf",
    "$ / sf",
    "per sf",
    "psf",
    "/sf",
    "amount/sf",
    "expense/sf",
)


def _norm(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip().casefold())


def _number(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        result = float(value)
        return result if math.isfinite(result) else None
    cleaned = re.sub(r"[$,%]", "", str(value)).replace(",", "").strip()
    cleaned = re.sub(r"\(([^)]+)\)", r"-\1", cleaned)
    try:
        result = float(cleaned)
    except ValueError:
        return None
    # Reject non-finite values ("nan", "inf", "Infinity"): Python's float()
    # accepts these strings, but a NaN/Inf rent or expense amount silently
    # corrupts dedupe, arithmetic reconciliation, and staged-JSON output
    # (json.dump emits bare NaN/Infinity, which is invalid JSON). Degrade
    # such a cell to "missing" so review catches it instead.
    return result if math.isfinite(result) else None


def _worksheet_max_row(ws, fallback=40):
    max_row = getattr(ws, "max_row", None)
    if isinstance(max_row, int) and max_row > 0:
        return max_row
    calculate_dimension = getattr(ws, "calculate_dimension", None)
    if callable(calculate_dimension):
        try:
            dimension = calculate_dimension(force=True)
        except TypeError:
            dimension = calculate_dimension()
        except Exception:
            dimension = None
        if dimension and dimension != "A1:A1":
            matches = re.findall(r"\d+", dimension)
            if matches:
                return max(int(match) for match in matches)
    return fallback


def _date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    text = str(value).strip()
    for fmt in (
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%m/%d/%y",
        "%m %d %Y",
        "%m.%d.%Y",
        "%B %d, %Y",
        "%b %d, %Y",
        "%B %Y",
        "%b %Y",
    ):
        try:
            return datetime.datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return text


def _field_for_header(header, synonyms):
    text = _norm(header)
    if not text:
        return None
    for field, variants in synonyms.items():
        if text in variants:
            return field
    for field, variants in synonyms.items():
        if any(
            len(variant) >= 4 and (variant in text or text in variant)
            for variant in variants
        ):
            return field
    return None


def _find_header(ws, synonyms, minimum_fields):
    best = None
    max_row = _worksheet_max_row(ws)
    for row_number, row in enumerate(
        ws.iter_rows(min_row=1, max_row=min(max_row, 40), values_only=True),
        start=1,
    ):
        mapping = {}
        claimed = set()
        for column, value in enumerate(row):
            field = _field_for_header(value, synonyms)
            if field and field not in claimed:
                mapping[column] = field
                claimed.add(field)
        if len(mapping) >= minimum_fields:
            return row_number, mapping
        if best is None or len(mapping) > len(best[1]):
            best = (row_number, mapping)
    return best if best and len(best[1]) >= minimum_fields else (None, {})


def _as_of_date(ws):
    max_row = _worksheet_max_row(ws, 12)
    for row in ws.iter_rows(
        min_row=1,
        max_row=min(max_row, 12),
        values_only=True,
    ):
        text = " ".join(str(value) for value in row if value is not None)
        match = re.search(
            r"(?:as of|rent roll date|effective date)[:\s]+"
            r"([A-Za-z]+\s+\d{1,2},?\s+\d{4}|\d{1,2}/\d{1,2}/\d{2,4})",
            text,
            re.IGNORECASE,
        )
        if match:
            return _date(match.group(1))
    for text in (getattr(ws, "title", ""),):
        month_year = re.search(
            r"\b("
            r"January|February|March|April|May|June|July|August|"
            r"September|October|November|December|"
            r"Jan|Feb|Mar|Apr|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec"
            r")\s+(\d{4})\b",
            text,
            re.IGNORECASE,
        )
        if month_year:
            return _date(month_year.group(0))
    return None


def _coerce_rent(field, value):
    if field in RENT_NUMBERS:
        return _number(value)
    if field in RENT_DATES:
        return _date(value)
    return str(value).strip() if value is not None else None


def _prepare_rent_data(data, confidence):
    first_name = data.get("tenant_first_name")
    last_name = data.get("tenant_last_name")
    if not data.get("tenant_name") and (first_name or last_name):
        parts = [part for part in (first_name, last_name) if part]
        data["tenant_name"] = " ".join(str(part).strip() for part in parts)
        confidence["tenant_name"] = "medium"
    if data.get("tenant_use"):
        lowered = _norm(data["tenant_use"])
        if lowered in {"vacant", "vacancy", "occupied"}:
            data.setdefault("occupancy_status", data["tenant_use"])
            confidence.setdefault("occupancy_status", "medium")
            data.pop("tenant_use", None)
            confidence.pop("tenant_use", None)
    for field in ("discounts", "balance"):
        if data.get(field) is not None and "reimbursement_structure" not in data:
            label = field.replace("_", " ")
            data["reimbursement_structure"] = f"{label}: {data[field]}"
            confidence["reimbursement_structure"] = "low"
    return data, confidence


def _coerce_expense(field, value):
    if field in EXPENSE_NUMBERS:
        number = _number(value)
        return int(number) if field == "period_year" and number is not None else number
    return str(value).strip() if value is not None else None


def _is_total_category(category):
    return _norm(category).startswith(TOTAL_PREFIXES)


def _extract_rows(ws, header_row, mapping, kind, source_path):
    records = []
    as_of_date = _as_of_date(ws) if kind == "rent_roll" else None
    for row_number, row in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        data = {}
        confidence = {}
        for column, field in mapping.items():
            if column >= len(row) or row[column] in (None, ""):
                continue
            value = (
                _coerce_rent(field, row[column])
                if kind == "rent_roll"
                else _coerce_expense(field, row[column])
            )
            if value is not None:
                data[field] = value
                confidence[field] = "high"
        if kind == "rent_roll":
            data, confidence = _prepare_rent_data(data, confidence)
            anchor = data.get("tenant_name") or data.get("unit_id") or data.get("suite")
            if not anchor:
                continue
            label = _norm(anchor)
            if label.startswith(("total", "subtotal", "average", "grand total")):
                continue
            if as_of_date:
                data["as_of_date"] = as_of_date
                confidence["as_of_date"] = "medium"
        else:
            category = data.get("category")
            if not category or data.get("amount") is None:
                continue
            if _is_total_category(category):
                continue
        records.append({
            "data": data,
            "confidence": confidence,
            "source": str(source_path),
            "sheet": ws.title,
            "source_row": row_number,
        })
    return records


def _left_fill(row):
    filled = []
    last = None
    for value in row:
        if value not in (None, ""):
            last = value
        filled.append(last)
    return filled


def _period_type_from_text(text):
    normalized = _norm(text).replace("-", " ")
    if "budget" in normalized:
        return "budget", "high"
    if "pro forma" in normalized or "proforma" in normalized:
        return "proforma", "high"
    if "projected" in normalized or "projection" in normalized:
        return "proforma", "medium"
    if "forecast" in normalized:
        return "forecast", "medium"
    if "stabilized" in normalized or "stabilised" in normalized:
        return "stabilized", "medium"
    if "actual" in normalized or "audited" in normalized or "historical" in normalized:
        return "actual", "high"
    return None, None


def _parse_period_header(*parts):
    text = " ".join(str(part) for part in parts if part not in (None, "")).strip()
    if not text:
        return None
    match = re.search(r"\b(19|20)\d{2}\b", text)
    if not match:
        return None
    normalized = _norm(text)
    value_field = (
        "amount_per_sf"
        if any(marker in normalized for marker in PER_SF_MARKERS)
        else "amount"
    )
    period_type, period_confidence = _period_type_from_text(text)
    return {
        "period_year": int(match.group(0)),
        "period_type": period_type,
        "period_type_confidence": period_confidence,
        "value_field": value_field,
    }


def _candidate_period_header(row, previous_row, column):
    current = row[column] if column < len(row) else None
    previous = previous_row[column] if previous_row and column < len(previous_row) else None
    if previous not in (None, "") and current not in (None, ""):
        parsed = _parse_period_header(previous, current)
        if parsed:
            return parsed
    parsed = _parse_period_header(current)
    if parsed:
        return parsed
    return _parse_period_header(previous)


def _find_wide_expense_header(ws):
    max_row = _worksheet_max_row(ws)
    rows = list(
        ws.iter_rows(
            min_row=1,
            max_row=min(max_row, 40),
            values_only=True,
        )
    )
    for index, row in enumerate(rows):
        previous_row = _left_fill(rows[index - 1]) if index else []
        category_columns = [
            column
            for column, value in enumerate(row)
            if _field_for_header(value, EXPENSE_SYNONYMS) == "category"
        ]
        for category_column in category_columns:
            period_columns = {}
            notes_column = None
            for column, value in enumerate(row):
                if column == category_column:
                    continue
                field = _field_for_header(value, EXPENSE_SYNONYMS)
                if field == "notes":
                    notes_column = column
                    continue
                parsed = _candidate_period_header(row, previous_row, column)
                if parsed:
                    period_columns[column] = parsed
            amount_columns = [
                column
                for column, parsed in period_columns.items()
                if parsed["value_field"] == "amount"
            ]
            if amount_columns:
                return {
                    "header_row": index + 1,
                    "category_column": category_column,
                    "period_columns": period_columns,
                    "notes_column": notes_column,
                }
    return None


def _extract_wide_expense_rows(ws, header, source_path):
    records = []
    header_row = header["header_row"]
    category_column = header["category_column"]
    period_columns = header["period_columns"]
    notes_column = header.get("notes_column")
    for row_number, row in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        category = (
            str(row[category_column]).strip()
            if category_column < len(row) and row[category_column] not in (None, "")
            else None
        )
        if not category or _is_total_category(category):
            continue
        notes = (
            str(row[notes_column]).strip()
            if notes_column is not None
            and notes_column < len(row)
            and row[notes_column] not in (None, "")
            else None
        )
        grouped = {}
        source_columns = {}
        for column, parsed in period_columns.items():
            if column >= len(row) or row[column] in (None, ""):
                continue
            value = _coerce_expense(parsed["value_field"], row[column])
            if value is None:
                continue
            key = (parsed["period_year"], parsed.get("period_type"))
            group = grouped.setdefault(
                key,
                {
                    "category": category,
                    "period_year": parsed["period_year"],
                },
            )
            group[parsed["value_field"]] = value
            if parsed.get("period_type"):
                group["period_type"] = parsed["period_type"]
            if notes:
                group["notes"] = notes
            source_columns.setdefault(key, []).append(column + 1)
        for key, data in grouped.items():
            if data.get("amount") is None:
                continue
            confidence = {
                "category": "high",
                "period_year": "high",
                "amount": "high",
            }
            if data.get("period_type"):
                confidence["period_type"] = (
                    period_columns[source_columns[key][0] - 1].get(
                        "period_type_confidence"
                    )
                    or "medium"
                )
            if data.get("amount_per_sf") is not None:
                confidence["amount_per_sf"] = "medium"
            if data.get("notes"):
                confidence["notes"] = "medium"
            first_col = min(source_columns[key])
            last_col = max(source_columns[key])
            locator = f"worksheet:{ws.title}:row:{row_number}:cols:{first_col}-{last_col}"
            records.append({
                "data": data,
                "confidence": confidence,
                "source": str(source_path),
                "sheet": ws.title,
                "source_row": row_number,
                "provenance": {
                    "source_locator": locator,
                    "layout": "wide_operating_statement",
                },
            })
    return records


def _record_dedupe_key(record, fields):
    data = record.get("data", {})
    return tuple(_norm(data.get(field)) for field in fields)


def _alternate_provenance(record, metadata_cache):
    source_path = record.get("source")
    provenance = {
        "source_path": source_path,
        "source_locator": (
            f"worksheet:{record.get('sheet')}:row:{record.get('source_row')}"
            if record.get("sheet") and record.get("source_row")
            else None
        ),
    }
    if source_path and Path(source_path).is_file():
        metadata = metadata_cache.get(source_path)
        if metadata is None:
            metadata = source_metadata(source_path)
            metadata_cache[source_path] = metadata
        provenance.update(metadata)
    return provenance


def _dedupe_records(records, fields):
    deduped = []
    by_key = {}
    metadata_cache = {}
    for record in records:
        key = _record_dedupe_key(record, fields)
        if not any(key):
            deduped.append(record)
            continue
        existing = by_key.get(key)
        if existing:
            existing.setdefault("alternate_provenance", []).append(
                _alternate_provenance(record, metadata_cache)
            )
            continue
        by_key[key] = record
        deduped.append(record)
    return deduped


def extract_financial_workbook(path):
    """Extract normalized long-form rent-roll and expense rows from XLSX."""
    path = Path(path)
    result = {
        "rent_roll_entries": [],
        "expense_records": [],
        "warnings": [],
    }
    try:
        workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        result["warnings"].append(f"Could not open {path.name}: {exc}")
        return result

    try:
        for ws in workbook.worksheets:
            rent_header, rent_map = _find_header(ws, RENT_ROLL_SYNONYMS, 3)
            expense_header, expense_map = _find_header(ws, EXPENSE_SYNONYMS, 3)
            wide_expense_header = _find_wide_expense_header(ws)
            sheet_name = _norm(ws.title)
            prefer_expense = "expense" in sheet_name
            normalized_expense_records = []
            if (
                expense_header
                and "amount" in expense_map.values()
                and (prefer_expense or not rent_header)
            ):
                normalized_expense_records = _extract_rows(
                    ws,
                    expense_header,
                    expense_map,
                    "expense",
                    path,
                )
            if normalized_expense_records:
                result["expense_records"].extend(normalized_expense_records)
            elif wide_expense_header and (prefer_expense or not rent_header):
                result["expense_records"].extend(
                    _extract_wide_expense_rows(
                        ws,
                        wide_expense_header,
                        path,
                    )
                )
            elif rent_header:
                result["rent_roll_entries"].extend(
                    _extract_rows(
                        ws,
                        rent_header,
                        rent_map,
                        "rent_roll",
                        path,
                    )
                )
    finally:
        workbook.close()
    result["rent_roll_entries"] = _dedupe_records(
        result["rent_roll_entries"],
        (
            "as_of_date",
            "unit_id",
            "suite",
            "tenant_name",
            "tenant_use",
            "lease_start",
            "lease_end",
            "monthly_rent",
            "annual_rent",
            "sf_leased",
        ),
    )
    result["expense_records"] = _dedupe_records(
        result["expense_records"],
        (
            "period_year",
            "period_type",
            "category",
            "amount",
            "amount_per_sf",
        ),
    )
    if not result["rent_roll_entries"] and not result["expense_records"]:
        result["warnings"].append(
            f"  [{path.name}] No normalized rent-roll or expense table found."
        )
    return result
