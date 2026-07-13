"""Search and export services for reviewed canonical comparable sales."""

import csv
import json
import shutil
from pathlib import Path

import openpyxl

from comp_builder import COMP_COLUMNS, col_letter_to_idx
from db import search_sale_comps


def _currency(value, decimals=0):
    if value is None:
        return ""
    return f"${float(value):,.{decimals}f}"


def _number(value, suffix=""):
    if value is None:
        return ""
    number = float(value)
    text = f"{number:,.0f}" if number.is_integer() else f"{number:,.2f}"
    return f"{text}{suffix}"


def _percentage(value):
    if value is None:
        return ""
    return f"{float(value):.2%}"


def sale_comp_to_report_fields(record, comp_number):
    """Map one database search row into the report comp-data contract."""
    city_line = ", ".join(
        value
        for value in (
            record.get("address_city"),
            record.get("address_state"),
            record.get("address_zip"),
        )
        if value
    )
    return {
        "COMP_NO": f"Sale No. {comp_number}",
        "COMP_SUBMARKET": record.get("submarket") or "",
        "COMP_ADDRESS_LINE1": record.get("address_street") or "",
        "COMP_ADDRESS_LINE2": city_line,
        "COMP_PROPERTY_TYPE": record.get("property_subtype")
        or record.get("property_type")
        or "",
        "COMP_SALE_PRICE": _currency(record.get("sale_price")),
        "COMP_SALE_DATE": record.get("sale_date") or "",
        "COMP_GBA_SF": _number(record.get("gba_sf"), " SF"),
        "COMP_PRICE_SF": _currency(record.get("price_per_sf"), decimals=2),
        "COMP_CAP_RATE": _percentage(record.get("cap_rate")),
        "COMP_YEAR_BUILT": _number(record.get("year_built")),
        "COMP_SITE_AREA": _number(record.get("site_area_sf"), " SF"),
        "COMP_STORIES": _number(record.get("stories")),
        "COMP_CONSTRUCTION": record.get("construction_type") or "",
        "COMP_CONDITION": record.get("condition") or "",
        "COMP_ZONING": record.get("zoning") or "",
        "COMP_FLOOD_ZONE": record.get("flood_zone") or "",
        "COMP_GRANTOR": record.get("grantor") or "",
        "COMP_GRANTEE": record.get("grantee") or "",
        "COMP_DEED_REF": record.get("deed_ref") or "",
        "COMP_VERIFICATION": record.get("verification_source") or "",
        "COMP_NOI": _currency(record.get("noi")),
        "COMP_NOI_SF": _currency(record.get("noi_per_sf"), decimals=2),
    }


def _mark_formula_cache_stale_if_live_workbook(output_workbook):
    """If *output_workbook* is a live assignment's workbook.xlsx (an
    `.axiom.json` state file lives next to it), record the same
    formula_cache_stale marker axiom.py's `_run_dilmore_calc` writes after
    any real save that touches that workbook.

    This export re-saves the workbook exactly like a real Dilmore write
    does -- wiping every OTHER cached formula result workbook-wide, since
    openpyxl has no formula engine -- but until now didn't set the marker
    that tells `validate`/`deliver` the cache needs a real Excel
    recalculation before the next delivery. That silently defeated round-3
    hardening's stale-cache guard (round-4 Fable adversarial review finding
    Q2). Deliberately checks the exact filename "workbook.xlsx", not just
    "an .axiom.json exists nearby" -- exporting to some OTHER file inside an
    assignment folder (not the live workbook `deliver` actually reads) must
    not mark that assignment's cache stale.
    """
    output_workbook = Path(output_workbook)
    if output_workbook.name != "workbook.xlsx":
        return
    state_file = output_workbook.parent / ".axiom.json"
    if not state_file.exists():
        return
    try:
        with open(state_file, encoding="utf-8") as f:
            state = json.load(f)
    except (OSError, json.JSONDecodeError):
        state = {}
    state["formula_cache_stale"] = True
    state["formula_cache_stale_mtime"] = output_workbook.stat().st_mtime
    tmp_file = output_workbook.parent / ".axiom.json.tmp"
    with open(tmp_file, "w", encoding="utf-8") as f:
        json.dump(state, f, indent=2)
    tmp_file.replace(state_file)


def export_sale_comps_to_workbook(
    records,
    source_workbook,
    output_workbook,
):
    """Write reviewed search results into a copied workbook's comp_data sheet."""
    source_workbook = Path(source_workbook)
    output_workbook = Path(output_workbook)
    output_workbook.parent.mkdir(parents=True, exist_ok=True)
    if source_workbook.resolve() != output_workbook.resolve():
        shutil.copy2(source_workbook, output_workbook)

    workbook = openpyxl.load_workbook(output_workbook)
    try:
        if "comp_data" not in workbook.sheetnames:
            raise ValueError("Workbook has no comp_data sheet.")
        sheet = workbook["comp_data"]
        for row in sheet.iter_rows(
            min_row=3,
            max_row=max(sheet.max_row, 3),
            min_col=1,
            max_col=max(col_letter_to_idx(col) for col in COMP_COLUMNS.values()),
        ):
            for cell in row:
                cell.value = None
        for row_number, record in enumerate(records, start=3):
            fields = sale_comp_to_report_fields(record, row_number - 2)
            for key, column in COMP_COLUMNS.items():
                sheet.cell(
                    row=row_number,
                    column=col_letter_to_idx(column),
                ).value = fields.get(key, "")
        workbook.save(output_workbook)
    finally:
        workbook.close()
    _mark_formula_cache_stale_if_live_workbook(output_workbook)
    return output_workbook


def export_sale_comps_csv(records, output_path):
    """Write canonical database search rows to a portable CSV."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    rows = list(records)
    fieldnames = sorted({key for row in rows for key in row})
    with open(output_path, "w", newline="", encoding="utf-8-sig") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    return output_path


__all__ = [
    "search_sale_comps",
    "sale_comp_to_report_fields",
    "export_sale_comps_to_workbook",
    "export_sale_comps_csv",
]
