"""
axiom.py — Axiom Commercial Appraisal Platform
================================================
The single entry point for all assignment workflow operations.

Commands
--------
  python axiom.py new <file_no> <client_name>   Create a new assignment
  python axiom.py engage <file_no>              Generate engagement documents
  python axiom.py deliver <file_no>             Generate delivery documents
  python axiom.py list                          Show all assignments
  python axiom.py status <file_no>              Show assignment details
  python axiom.py dashboard                     Build a visual HTML dashboard

Setup
-----
Place axiom.py, fill_engine.py, and config.json together in one folder.
That folder should contain:
  templates/    — read-only source templates (Word docs + workbook)
  assignments/  — auto-created; one subfolder per assignment

Requirements: pip install python-docx openpyxl
"""

import sys
import json
import shutil
import datetime
import html as html_lib
import hashlib
import re
import tempfile
from pathlib import Path

from field_registry import (
    audit_assignment_contract,
    load_registry,
    registry_version,
)
from fill_engine import fill_document, load_variables
from comp_builder import inject_comp_section
from adjustment_grid import (
    MAX_DATA_ROW as GRID_MAX_DATA_ROW,
    inject_all_adjustment_grids,
    read_header_map,
)
from dilmore import dilmore_factor, dilmore_adj_pct, dilmore_summary
from media_blocks import create_media_directories, inject_media_blocks
from structured_blocks import inject_ownership_history
from validation import (
    find_docx_placeholders,
    intake_json_findings,
    validate_assignment,
)


# ─── Bootstrap ────────────────────────────────────────────────────────────────

BASE_DIR = Path(__file__).parent

with open(BASE_DIR / "config.json", encoding="utf-8") as _f:
    CONFIG = json.load(_f)

TEMPLATES_DIR   = BASE_DIR / CONFIG["templates_dir"]
ASSIGNMENTS_DIR = BASE_DIR / CONFIG["assignments_dir"]
WORKBOOK_TPL    = BASE_DIR / CONFIG["workbook_template"]
FIELD_REGISTRY  = BASE_DIR / CONFIG["field_registry"]
APP_VERSION     = CONFIG.get("app_version", "unversioned")
STAGES          = CONFIG["stages"]
try:
    OPTIONAL_BLANK_FIELDS = frozenset(
        key
        for key, definition in load_registry(FIELD_REGISTRY)["fields"].items()
        if definition.get("required") is False
    )
except Exception:
    OPTIONAL_BLANK_FIELDS = frozenset()

ASSIGNMENTS_DIR.mkdir(exist_ok=True)


# ─── Assignment helpers ───────────────────────────────────────────────────────

def _find_assignment(file_no):
    """Return the assignment directory for a given file number, or None."""
    for d in ASSIGNMENTS_DIR.iterdir():
        if d.is_dir() and d.name.startswith(f"{file_no}_"):
            return d
    return None


FILE_NO_PATTERN = re.compile(
    r"^[A-Za-z0-9](?:[A-Za-z0-9._-]{0,62}[A-Za-z0-9])?$"
)


def _valid_file_no(file_no):
    return bool(
        FILE_NO_PATTERN.fullmatch(str(file_no))
        and ".." not in str(file_no)
    )


def _safe_client_component(client_name):
    safe = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "-", str(client_name))
    safe = re.sub(r"\s+", "_", safe.strip())
    safe = re.sub(r"\.{2,}", "-", safe)
    safe = re.sub(r"[-_]{2,}", "_", safe)
    return safe.strip(" ._-")[:100]


def _load_state(adir):
    state_file = adir / ".axiom.json"
    if state_file.exists():
        with open(state_file, encoding="utf-8") as f:
            return json.load(f)
    return {}


def _save_state(adir, state):
    # Atomic write (write-to-temp then rename) so a process interrupted
    # mid-write (crash, kill, disk full) can never leave .axiom.json half
    # written -- a partial/corrupt JSON file used to fail-open every guard
    # that reads this state (e.g. validation.py's stale-Dilmore-cache check),
    # since _load_state would either raise or silently see a truncated dict
    # missing the very flags meant to protect against a bad delivery (round-4
    # Fable adversarial review finding Q8).
    state_file = adir / ".axiom.json"
    tmp_file = adir / ".axiom.json.tmp"
    with open(tmp_file, "w", encoding="utf-8") as f:
        json.dump(state, f, indent=2)
    try:
        tmp_file.replace(state_file)
    except OSError:
        # The atomic rename itself can fail on some environments (a sync
        # client or antivirus holding a lock, a locked-output scenario that
        # also happens to affect the state file) -- this call is frequently
        # made from inside an `except` block trying to record why some other
        # operation just failed, so letting the rename's own failure escape
        # uncaught would replace a clear error message with a raw traceback.
        # Fall back to a direct (non-atomic) write rather than lose the
        # state update entirely.
        with open(state_file, "w", encoding="utf-8") as f:
            json.dump(state, f, indent=2)
        tmp_file.unlink(missing_ok=True)


def _find_json(adir):
    """Find the variables JSON file in an assignment directory."""
    matches = sorted(adir.glob("*_variables.json"))
    return matches[0] if len(matches) == 1 else None


def _today():
    return datetime.date.today().strftime("%B %d, %Y")


# ─── Commands ─────────────────────────────────────────────────────────────────

def cmd_new(args):
    """Create a new assignment folder and copy in the workbook template."""
    if len(args) < 2:
        print("Usage: python axiom.py new <file_no> <client_name>")
        print("Example: python axiom.py new DEMO-001 Northstar Example Holdings")
        return

    file_no     = args[0]
    client_name = " ".join(args[1:])
    if not _valid_file_no(file_no):
        print(
            "  Invalid file number. Use 1-64 letters, numbers, dots, "
            "hyphens, or underscores; path-like values are not allowed."
        )
        return False
    safe_client = _safe_client_component(client_name)
    if not safe_client:
        print("  Client name has no usable filename characters.")
        return False
    folder_name = f"{file_no}_{safe_client}"
    adir        = ASSIGNMENTS_DIR / folder_name

    if adir.exists():
        print(f"  Assignment {file_no} already exists at:")
        print(f"  {adir}")
        return

    # Create folder structure
    adir.mkdir(parents=True)
    (adir / "outputs").mkdir()
    create_media_directories(adir)

    # Copy workbook template
    if WORKBOOK_TPL.exists():
        shutil.copy(WORKBOOK_TPL, adir / "workbook.xlsx")
        print(f"  Copied workbook template")
    else:
        print(f"  Warning: workbook template not found at {WORKBOOK_TPL}")

    # Write state file
    _save_state(adir, {
        "file_no":   file_no,
        "client":    client_name,
        "stage":     "new",
        "created":   _today(),
        "engaged":   None,
        "delivered": None,
        "app_version": APP_VERSION,
        "schema_version": registry_version(FIELD_REGISTRY),
    })

    print(f"\n  Created: {folder_name}")
    print(f"  Location: {adir}")
    print(f"\n  Next steps:")
    print(f"    1. Open workbook.xlsx and fill out the Intake tab")
    print(f"    2. Click the Export JSON button")
    print(f"    3. Run: python axiom.py engage {file_no}")


def cmd_engage(args):
    """Generate engagement-stage documents (engagement letter + doc request)."""
    if not args:
        print("Usage: python axiom.py engage <file_no>")
        return

    file_no = args[0]
    adir    = _find_assignment(file_no)

    if not adir:
        print(f"  Assignment {file_no} not found.")
        print(f"  Run: python axiom.py new {file_no} <client_name>")
        return

    # Check for JSON
    json_path = _find_json(adir)
    if not json_path:
        print(f"  No variables.json found in {adir.name}")
        print(f"  Fill out the Intake tab in workbook.xlsx and click Export JSON first.")
        return

    workbook_path = adir / "workbook.xlsx"
    if workbook_path.exists():
        try:
            stale_fields, freshness_warnings = intake_json_findings(
                workbook_path,
                json_path,
                FIELD_REGISTRY,
            )
        except Exception as exc:
            print(f"  Intake/JSON freshness check failed: {exc}")
            return False
        for warning in freshness_warnings:
            print(f"  Warning: {warning}")
        if stale_fields:
            print(
                f"  Engagement blocked: {len(stale_fields)} canonical Intake "
                "field(s) differ from exported JSON."
            )
            for key in stale_fields:
                print(f"    - {key}")
            print("  Re-export JSON from the current workbook, then retry.")
            return False

    print(f"\n  Loading variables from {json_path.name} ...")
    try:
        variables = load_variables(json_path=json_path)
    except Exception as exc:
        state = _load_state(adir)
        state["last_engagement_status"] = "input_failed"
        state["last_engagement_error"] = str(exc)
        _save_state(adir, state)
        print(f"  Engagement input loading failed: {exc}")
        return False
    print(f"  {len(variables)} variables loaded")

    outputs_dir = adir / "outputs"
    outputs_dir.mkdir(exist_ok=True)
    stage_cfg   = STAGES["engage"]

    print(f"\n  Generating engagement documents:")
    pending_outputs = []
    for doc_cfg in stage_cfg["documents"]:
        template_path = TEMPLATES_DIR / doc_cfg["template"]
        output_name   = f"{file_no}_{doc_cfg['output']}"
        output_path   = outputs_dir / output_name

        if not template_path.exists():
            error = f"Template not found: {doc_cfg['template']}"
            for working_path, _ in pending_outputs:
                working_path.unlink(missing_ok=True)
            state = _load_state(adir)
            state["last_engagement_status"] = "generation_failed"
            state["last_engagement_error"] = error
            _save_state(adir, state)
            print(f"    {error}")
            return False

        with tempfile.NamedTemporaryFile(
            dir=outputs_dir,
            prefix=f".{output_name}.",
            suffix=".tmp",
            delete=False,
        ) as temp_output:
            working_path = Path(temp_output.name)
        try:
            fill_document(
                template_path,
                working_path,
                variables,
                optional_blank_keys=OPTIONAL_BLANK_FIELDS,
            )
        except Exception as exc:
            working_path.unlink(missing_ok=True)
            for pending_path, _ in pending_outputs:
                pending_path.unlink(missing_ok=True)
            state = _load_state(adir)
            state["last_engagement_status"] = "generation_failed"
            state["last_engagement_error"] = str(exc)
            _save_state(adir, state)
            print(f"    Engagement generation failed: {exc}")
            return False
        pending_outputs.append((working_path, output_path))

    if not pending_outputs:
        state = _load_state(adir)
        state["last_engagement_status"] = "generation_failed"
        state["last_engagement_error"] = "No engagement documents configured."
        _save_state(adir, state)
        print("  No engagement documents were generated.")
        return False

    try:
        for working_path, output_path in pending_outputs:
            working_path.replace(output_path)
            print(f"    OK: {output_path.name}")
    except Exception as exc:
        for working_path, _ in pending_outputs:
            working_path.unlink(missing_ok=True)
        state = _load_state(adir)
        state["last_engagement_status"] = "generation_failed"
        state["last_engagement_error"] = str(exc)
        _save_state(adir, state)
        print(f"    Engagement output replacement failed: {exc}")
        return False

    # Update state
    state = _load_state(adir)
    state["stage"]   = "engaged"
    state["engaged"] = _today()
    state["last_engagement_status"] = "completed"
    state.pop("last_engagement_error", None)
    _save_state(adir, state)

    print(f"\n  {len(pending_outputs)} document(s) ready in:")
    print(f"  {outputs_dir}")
    return True


def _inject_all_narratives(doc_path, workbook_path, variables):
    """
    Replace all [[X_NARRATIVE]] and [[X_OVERVIEW]] placeholders in a filled
    Word document with AI-generated USPAP-compliant prose.
    Delegates to narrative_generator.inject_all_narratives().
    """
    try:
        from narrative_generator import inject_all_narratives
    except ImportError as e:
        print(f"    Warning: narrative_generator import failed — {e}")
        return

    print(f"\n  Generating AI narratives ...")
    try:
        results = inject_all_narratives(doc_path, workbook_path, variables)
        if not results:
            print(f"    (no narrative placeholders found in document)")
    except Exception as e:
        print(f"    Warning: narrative injection error — {e}")


def cmd_deliver(args):
    """Generate a validated final report, or an explicitly incomplete draft."""
    if not args:
        print("Usage: python axiom.py deliver <file_no> [--draft]")
        return

    file_no = args[0]
    options = set(args[1:])
    unknown_options = options - {"--draft"}
    if unknown_options:
        print(f"  Unknown deliver option(s): {', '.join(sorted(unknown_options))}")
        print("Usage: python axiom.py deliver <file_no> [--draft]")
        return False
    draft_mode = "--draft" in options

    adir    = _find_assignment(file_no)

    if not adir:
        print(f"  Assignment {file_no} not found.")
        return False

    validation = check_delivery_readiness(adir)
    if not validation["ready"] and not draft_mode:
        blocker_count = (
            len(validation.get("errors", []))
            + len(validation.get("missing", []))
            + len(validation.get("unresolved_blocks", {}))
        )
        state = _load_state(adir)
        state["last_delivery_attempt"] = _today()
        state["last_delivery_status"] = "blocked"
        state["last_delivery_blocker_count"] = blocker_count
        state.pop("last_delivery_error", None)
        _save_state(adir, state)

        print(f"\n  Delivery blocked: {blocker_count} validation issue(s).")
        print(f"  Run: python axiom.py validate {file_no}")
        print(
            f"  To generate an explicitly incomplete report without changing "
            f"delivery state:"
        )
        print(f"    python axiom.py deliver {file_no} --draft")
        return False

    if draft_mode and not validation["ready"]:
        print("\n  DRAFT MODE: validation issues will remain visible in the report.")
        print("  Assignment delivery state will not be changed.")

    json_path     = _find_json(adir)
    workbook_path = adir / "workbook.xlsx"

    if not json_path:
        print(f"  No variables.json found. Export from Intake tab first.")
        return False

    if not workbook_path.exists():
        print(f"  workbook.xlsx not found in {adir.name}")
        return False

    # Auto-run the Dilmore size-adjustment calc before generating anything,
    # so Derek never has to remember a separate `dilmore` step after editing
    # a comp's GBA. Per Derek's explicit choice (2026-07-10): keep this as
    # a tested Python calculation rather than a live Excel formula -- see
    # docs/ADJUSTMENT_GRID_DESIGN.md. A missing adjustment-grid tab (older
    # in-flight assignment), no comp GBAs yet entered, or values that are
    # already current are all routine and never block delivery. But if this
    # calc actually changes a Size Factor/Adj % value, its save() just
    # discarded every other cached formula result in the workbook (Net
    # Adjustment, Indicated Value, the outputs tab -- see _run_dilmore_calc's
    # docstring) -- there's no way to safely proceed to generate a report
    # from data_only=True reads of a workbook whose cache was just
    # invalidated. Stop and tell Derek to recalculate/save in Excel instead of
    # silently delivering a report with blank conclusion numbers.
    #
    # Skipped entirely in draft mode: this auto-run calls wb.save() on real
    # writes, which both mutates workbook.xlsx and (via _run_dilmore_calc's
    # new state write) sets .axiom.json's formula_cache_stale flag -- either
    # one contradicts the "Assignment delivery state will not be changed"
    # promise draft mode prints above, and a draft is meant to be a quick,
    # side-effect-free preview of an otherwise-incomplete assignment (Fable
    # adversarial review finding P4).
    if not draft_mode:
        dilmore_result = _run_dilmore_calc(workbook_path)
        if dilmore_result["ok"] and dilmore_result.get("count"):
            print(f"  [Dilmore] {dilmore_result['message']}")
            if dilmore_result.get("changed"):
                state = _load_state(adir)
                state["last_delivery_attempt"] = _today()
                state["last_delivery_status"] = "input_failed"
                state["last_delivery_blocker_count"] = 1
                state["last_delivery_error"] = (
                    f"Size adjustments in {dilmore_result['tab']} just changed. "
                    "openpyxl cannot recalculate Excel formulas, so every other "
                    "cached formula result in this workbook (Net Adjustment, "
                    "Indicated Value, the outputs tab, etc.) is now stale or "
                    "blank until Excel recalculates it. Open workbook.xlsx, let "
                    "it fully recalculate (e.g. press F9), save, then re-run "
                    "deliver."
                )
                _save_state(adir, state)
                print(f"\n  {state['last_delivery_error']}")
                return False
        elif not dilmore_result["ok"]:
            print(f"  [Dilmore] skipped -- {dilmore_result['message']}")
    else:
        # A generic skip line used to print unconditionally here regardless
        # of whether anything was actually pending -- Derek had no way to
        # tell "routine skip, nothing to do" from "this draft's Size Factor/
        # Adj % are genuinely incomplete" without separately running
        # `dilmore` or `validate`. _dilmore_staleness_warnings is read-only
        # (safe in draft mode) and is the same check `validate` already
        # surfaces, so reuse it here to make the skip message specific when
        # there's real pending work (round-4 Fable adversarial review
        # finding Q5).
        staleness_warnings = _dilmore_staleness_warnings(workbook_path)
        if staleness_warnings:
            print(
                "  [Dilmore] skipped in draft mode -- pending size "
                "adjustment(s) were NOT calculated, so this draft's Size "
                "Factor/Adj % (and anything derived from them) may be "
                "incomplete:"
            )
            for warning in staleness_warnings:
                print(f"    [!] {warning}")
        else:
            print("  [Dilmore] skipped in draft mode -- nothing pending.")

    print(f"\n  Loading variables ...")
    try:
        variables = load_variables(
            json_path=json_path,
            workbook_path=workbook_path,
        )
    except Exception as exc:
        state = _load_state(adir)
        state["last_delivery_attempt"] = _today()
        state["last_delivery_status"] = "input_failed"
        state["last_delivery_blocker_count"] = 1
        state["last_delivery_error"] = str(exc)
        _save_state(adir, state)
        print(f"  Delivery input loading failed: {exc}")
        return False
    print(f"  {len(variables)} variables loaded (JSON + workbook outputs tab)")

    outputs_dir = adir / "outputs"
    outputs_dir.mkdir(exist_ok=True)
    stage_cfg   = STAGES["deliver"]

    print(f"\n  Generating delivery documents:")
    count = 0
    for doc_cfg in stage_cfg["documents"]:
        template_path = TEMPLATES_DIR / doc_cfg["template"]
        draft_marker  = "DRAFT_" if draft_mode else ""
        output_name   = f"{file_no}_{draft_marker}{doc_cfg['output']}"
        output_path   = outputs_dir / output_name

        if not template_path.exists():
            print(f"    ✗  Template not found: {doc_cfg['template']}")
            continue

        with tempfile.NamedTemporaryFile(
            dir=outputs_dir,
            prefix=f".{output_name}.",
            suffix=".tmp",
            delete=False,
        ) as temp_output:
            working_path = Path(temp_output.name)

        try:
            state = _load_state(adir)
            state["delivery_provenance"] = {
                "app_version": APP_VERSION,
                "schema_version": registry_version(FIELD_REGISTRY),
                "template": doc_cfg["template"],
                "template_sha256": hashlib.sha256(
                    template_path.read_bytes()
                ).hexdigest(),
            }
            _save_state(adir, state)

            result = fill_document(
                template_path,
                working_path,
                variables,
                optional_blank_keys=OPTIONAL_BLANK_FIELDS,
            )
            if result.get("removed_sections"):
                print(
                    "       Removed sections: "
                    + ", ".join(result["removed_sections"])
                )
            if result["missing"]:
                print(
                    f"       {len(result['missing'])} unfilled keys: "
                    + ", ".join(result["missing"])
                )
            if result["blocks"]:
                print(
                    f"       {len(result['blocks'])} block placeholders "
                    "left intact: "
                    + ", ".join(result["blocks"])
                )

            if doc_cfg.get("inject_comps") and working_path.exists():
                comp_template = TEMPLATES_DIR / "comp_block_template.docx"
                if comp_template.exists():
                    print("\n  Injecting comp pages ...")
                    inject_comp_section(
                        working_path,
                        comp_template,
                        workbook_path,
                    )
                else:
                    print(
                        "    Warning: comp_block_template.docx not found "
                        "in templates/"
                    )

            if doc_cfg.get("inject_comps") and working_path.exists():
                print("\n  Injecting assignment media ...")
                media_results = inject_media_blocks(working_path, adir)
                if media_results:
                    for block, image_count in sorted(media_results.items()):
                        print(f"    OK: {block}: {image_count} image(s)")
                else:
                    print("    (no media assets injected)")

            if doc_cfg.get("inject_comps") and working_path.exists():
                if inject_ownership_history(working_path, variables):
                    print("    OK: OWNERSHIP_HISTORY_TABLE")

            if doc_cfg.get("inject_comps") and working_path.exists():
                grid_results = inject_all_adjustment_grids(
                    working_path, workbook_path
                )
                for block, row_count in sorted(grid_results.items()):
                    if row_count:
                        print(f"    OK: {block}: {row_count} row(s)")

            if doc_cfg.get("inject_comps") and working_path.exists():
                _inject_all_narratives(
                    working_path,
                    workbook_path,
                    variables,
                )

            working_path.replace(output_path)
            print(f"    OK: {output_name}")
            count += 1
        except Exception as exc:
            working_path.unlink(missing_ok=True)
            state = _load_state(adir)
            state["last_delivery_attempt"] = _today()
            state["last_delivery_status"] = "generation_failed"
            state["last_delivery_blocker_count"] = 1
            state["last_delivery_error"] = str(exc)
            _save_state(adir, state)
            print(f"    Generation failed for {output_name}: {exc}")
            return False

    if count == 0:
        state = _load_state(adir)
        state["last_delivery_attempt"] = _today()
        state["last_delivery_status"] = "generation_failed"
        state["last_delivery_blocker_count"] = 1
        state["last_delivery_error"] = "No delivery documents were generated."
        _save_state(adir, state)
        print("\n  No delivery documents were generated.")
        return False

    state = _load_state(adir)
    state["last_delivery_attempt"] = _today()

    if draft_mode:
        state["last_delivery_status"] = "draft_generated"
        # Deliberately NOT clearing last_delivery_error here: a prior real
        # (non-draft) delivery attempt may have left a genuinely still-true
        # guidance message here -- most importantly P1's "open workbook.xlsx,
        # let it recalculate, save, then re-run deliver" instruction, which
        # stays relevant until a real delivery actually resolves it. A draft
        # run used to silently erase that message even though the underlying
        # formula_cache_stale condition it describes was untouched (round-4
        # Fable adversarial review finding Q4).
        _save_state(adir, state)
        print(f"\n  Draft generated; assignment stage remains {state.get('stage', 'new')}.")
        print(f"  {count} document(s) ready for review in:")
        print(f"  {outputs_dir}")
        return True

    remaining_placeholders = []
    for doc_cfg in stage_cfg["documents"]:
        output_path = outputs_dir / f"{file_no}_{doc_cfg['output']}"
        if output_path.exists():
            remaining_placeholders.extend(find_docx_placeholders(output_path))

    remaining_placeholders = sorted(set(remaining_placeholders))
    if remaining_placeholders:
        state["last_delivery_status"] = "blocked_after_generation"
        state["last_delivery_blocker_count"] = len(remaining_placeholders)
        state.pop("last_delivery_error", None)
        _save_state(adir, state)
        print(
            f"\n  Delivery state NOT changed: generated output still contains "
            f"{len(remaining_placeholders)} placeholder(s)."
        )
        print(f"  Run: python axiom.py validate {file_no}")
        return False

    state["stage"] = "delivered"
    state["delivered"] = _today()
    state["last_delivery_status"] = "completed"
    state["last_delivery_blocker_count"] = 0
    state.pop("last_delivery_error", None)
    # A real (non-draft) delivery only ever reaches this point if
    # check_delivery_readiness/_dilmore_cache_still_stale already let it
    # through, meaning the cache was either never marked stale or was
    # resaved/recalculated since -- so it's safe, and correct, to clear the
    # marker now rather than leave it set forever after the condition it
    # described has actually been resolved (round-4 Fable adversarial
    # review finding Q8).
    state.pop("formula_cache_stale", None)
    state.pop("formula_cache_stale_mtime", None)
    _save_state(adir, state)

    print(f"\n  {count} document(s) ready in:")
    print(f"  {outputs_dir}")
    return True


# Adjustment-grid tab layouts cmd_dilmore/deliver's auto-run understand,
# newest first. sca_adjustment_grid is the current (Phase 6) tab; size_adj
# is the older tab an assignment engaged before Phase 6 shipped may still
# have (see docs/ADJUSTMENT_GRID_DESIGN.md, Open Question e -- graceful
# skip/detection rather than a forced migration). Both use the same
# 10-comp-row layout starting at row 7; only the column positions and a
# couple of header labels differ.
#
# Columns are identified by header text, resolved at runtime via
# adjustment_grid.read_header_map, not by fixed index -- this used to be a
# hardcoded {"comp_gba_col": 3, "factor_col": 11, "adj_pct_col": 12}-style
# dict, which silently wrote Size Factor/Adj % into whatever now occupied
# those column *positions* if a column was ever hand-inserted into the
# sheet, with no error (round-4 Fable adversarial review finding Q6).
# read_grid_rows (the report-reading side) was already header-driven by
# design; this makes the write side agree with it.
_DILMORE_TAB_LAYOUTS = {
    "sca_adjustment_grid": {
        "comp_gba_header": "Comp GBA (SF)",
        "factor_header": "Size Factor",
        "adj_pct_header": "Size Adj %",
    },
    "size_adj": {
        "comp_gba_header": "Comp GBA (SF)",
        "factor_header": "Size Factor",
        "adj_pct_header": "Adj %",
    },
}


def _run_dilmore_calc(workbook_path):
    """Core Dilmore size-adjustment calculation, shared by the standalone
    `dilmore` command and deliver's automatic pre-flight step.

    Detects whichever adjustment-grid tab the given workbook actually has
    and writes Size Factor / Adj % into the right columns for that layout.
    Adj % is written as a fraction (factor - 1), matching the true
    percentage cell format on both tabs -- not dilmore_adj_pct()'s
    percentage-point number, which would display 100x too large.

    Returns a dict, always with "ok" and "message" keys:
      {"ok": False, "message": "..."}                         -- hard error
      {"ok": True, "count": 0, "message": "..."}               -- nothing to do
      {"ok": True, "count": N, "tab": ..., "message": "...",
       "subject_gba": ..., "curve": ..., "summary": [...]}     -- wrote N rows
    """
    import openpyxl
    # Two handles on the same file: `wb` (data_only=False) is the only one
    # ever written to or saved, preserving every other sheet's live formula
    # objects. `wb_values` (data_only=True) is read-only and used for every
    # value this function needs to read -- curve, subject GBA, and each
    # comp's GBA. A comp's GBA entered as a live Excel formula (e.g.
    # referencing another cell) used to be read from `wb` alone, which hands
    # back the raw formula string, not its computed number; float() on that
    # string always raised, so the row was silently skipped with no Size
    # Factor ever written -- while the read-only staleness check in
    # `_dilmore_staleness_warnings` (already data_only=True) saw the correct
    # cached number and kept reporting "runs automatically on delivery" even
    # though it never actually would (round-4 Fable adversarial review
    # finding Q7). Reading every value through the same data_only=True view
    # the rest of the platform uses closes that gap.
    wb = openpyxl.load_workbook(str(workbook_path))
    wb_values = openpyxl.load_workbook(str(workbook_path), data_only=True)

    tab_name = next(
        (name for name in _DILMORE_TAB_LAYOUTS if name in wb.sheetnames), None
    )
    if tab_name is None:
        wb.close()
        wb_values.close()
        return {
            "ok": False,
            "message": (
                "No sca_adjustment_grid or size_adj tab found -- is this "
                "workbook from a compatible template?"
            ),
        }

    layout = _DILMORE_TAB_LAYOUTS[tab_name]
    sa = wb[tab_name]              # write target
    sa_values = wb_values[tab_name]  # read-only cached values

    header_map = read_header_map(sa)
    comp_gba_col = header_map.get(layout["comp_gba_header"])
    factor_col = header_map.get(layout["factor_header"])
    adj_pct_col = header_map.get(layout["adj_pct_header"])
    missing_headers = [
        name for name, col in (
            (layout["comp_gba_header"], comp_gba_col),
            (layout["factor_header"], factor_col),
            (layout["adj_pct_header"], adj_pct_col),
        )
        if col is None
    ]
    if missing_headers:
        wb.close()
        wb_values.close()
        return {
            "ok": False,
            "message": (
                f"{tab_name} is missing expected column header(s): "
                f"{', '.join(missing_headers)}. Check row {6} hasn't been "
                "restructured."
            ),
        }

    curve_val = sa_values["B3"].value
    try:
        curve = float(curve_val) if curve_val else 85.0
    except (ValueError, TypeError):
        curve = 85.0

    # Read subject GBA from Intake tab
    subject_gba = None
    if "Intake" in wb_values.sheetnames:
        intake = wb_values["Intake"]
        for row in intake.iter_rows(min_row=2, values_only=True):
            if row[0] and str(row[0]).strip().upper() == "GBA":
                try:
                    raw = str(row[1]).replace(",", "").replace(" SF", "").strip()
                    subject_gba = float(raw)
                except (ValueError, TypeError):
                    pass
                break

    if not subject_gba:
        json_path = _find_json(workbook_path.parent)
        if json_path:
            import json
            with open(json_path) as f:
                jdata = json.load(f)
            raw = jdata.get("GBA", "")
            try:
                subject_gba = float(str(raw).replace(",", "").replace(" SF", ""))
            except (ValueError, TypeError):
                pass

    if not subject_gba:
        wb.close()
        wb_values.close()
        return {
            "ok": False,
            "message": (
                "Could not find Subject GBA in Intake tab or variables.json. "
                "Enter GBA in the Intake tab (key: GBA) and re-export JSON."
            ),
        }

    # Read comp GBAs from the detected tab's comp rows, following the exact
    # same "Sale No." anchor convention read_grid_rows uses: for
    # sca_adjustment_grid, scan every row from 7 up through
    # GRID_MAX_DATA_ROW, stopping at the first row whose anchor doesn't
    # match. Rows 7-16 used to be scanned unconditionally with no anchor
    # check at all, which let this function write a Size Factor to a row
    # read_grid_rows would never treat as a comp (e.g. one whose "Sale
    # No." label got cleared/corrupted) -- Dilmore's own console output
    # ("Wrote N size adjustment(s)") would then disagree with what the
    # delivered report actually contained, reassuring Derek a comp was
    # processed when the report silently dropped it (Fable adversarial
    # review finding P3, the same root cause as adjustment_grid.py's P2
    # orphan-anchor gap). size_adj (the older, pre-Phase-6 tab) predates
    # the "Sale No." anchor convention entirely and keeps its original
    # fixed, unconditional 7-16 window.
    row_idxs = []
    comp_gbas = []
    if tab_name == "sca_adjustment_grid":
        scan_rows = []
        for candidate_row in range(7, GRID_MAX_DATA_ROW + 1):
            comp_label = sa_values.cell(row=candidate_row, column=1).value
            if not (isinstance(comp_label, str) and comp_label.startswith("Sale No.")):
                break
            scan_rows.append(candidate_row)
    else:
        scan_rows = list(range(7, 17))

    for row_idx in scan_rows:
        comp_gba_val = sa_values.cell(row=row_idx, column=comp_gba_col).value
        if not comp_gba_val:
            continue
        try:
            comp_gba = float(str(comp_gba_val).replace(',', '').replace(' SF', ''))
        except (ValueError, TypeError):
            continue
        row_idxs.append(row_idx)
        comp_gbas.append(comp_gba)

    if not comp_gbas:
        wb.close()
        wb_values.close()
        return {
            "ok": True,
            "count": 0,
            "changed": False,
            "tab": tab_name,
            "message": f"No comp GBAs found in {tab_name}.",
        }

    # dilmore_summary computes ratio = comp_gba / subject_gba (A_c/A_s) and
    # applies it against the requested curve for every comp in one call --
    # this used to be reimplemented inline here with the two raw GBAs passed
    # as separate positional args to dilmore_factor/dilmore_adj_pct (a 3-arg
    # call against their real 2-arg (ratio, curve) signature), which raised
    # TypeError on every real run and also had the ratio backwards.
    try:
        summary = dilmore_summary(subject_gba, comp_gbas, curve)
    except ValueError as exc:
        wb.close()
        wb_values.close()
        return {
            "ok": False,
            "message": (
                f"{exc} Fix the curve in {tab_name}!B3 to one of "
                f"80, 82.5, 85, 87.5, 90."
            ),
        }

    # Only actually write (and therefore save) if a value is really
    # changing. This matters a lot more than it looks: openpyxl has no
    # formula engine, so ANY load-modify-save cycle on this workbook
    # discards every OTHER formula cell's cached value workbook-wide, not
    # just the Size Factor/Adj % cells this function touches -- confirmed
    # directly this session (see HANDOFF.md, "Phase 6 completion"). Skipping
    # the save entirely when nothing changed means re-running `deliver` on
    # already-current data (the common case) never touches the file.
    changed = False
    for row_idx, row in zip(row_idxs, summary):
        new_factor = round(row["factor"], 4)
        new_adj_pct = round(row["factor"] - 1, 4)
        # Compare against the CACHED value (data_only=True) rather than
        # whatever `sa`'s data_only=False view happens to hold -- if these
        # cells were ever formulas instead of plain numbers, data_only=False
        # would show the formula string and always look "changed", forcing
        # an unnecessary save (and cache wipe) every single run.
        old_factor_cached = sa_values.cell(row=row_idx, column=factor_col).value
        old_adj_pct_cached = sa_values.cell(row=row_idx, column=adj_pct_col).value
        if old_factor_cached != new_factor or old_adj_pct_cached != new_adj_pct:
            changed = True
        factor_cell = sa.cell(row=row_idx, column=factor_col)
        adj_pct_cell = sa.cell(row=row_idx, column=adj_pct_col)
        factor_cell.value = new_factor
        adj_pct_cell.value = new_adj_pct

    if not changed:
        wb.close()
        wb_values.close()
        return {
            "ok": True,
            "count": len(row_idxs),
            "changed": False,
            "tab": tab_name,
            "message": (
                f"{len(row_idxs)} size adjustment(s) in {tab_name} already "
                "up to date."
            ),
            "subject_gba": subject_gba,
            "curve": curve,
            "summary": summary,
        }

    # A real write is happening, which means this save WILL blank every
    # other cached formula result in the workbook (Net Adjustment,
    # Indicated Value, the outputs tab -- everything), per the note above.
    # There is no Excel/LibreOffice automation in this codebase to
    # recalculate before a caller's next data_only=True read, so the caller
    # (cmd_deliver) must treat "changed": True as a hard stop rather than
    # proceeding to read/deliver against a workbook whose cache it just
    # invalidated.
    wb.save(str(workbook_path))

    # Record exactly when this happened and the workbook's mtime at that
    # moment, so validation.py's _dilmore_cache_still_stale can tell a
    # LATER delivery attempt (where Dilmore finds nothing new to write,
    # "changed": False) that the cache from THIS write was never actually
    # recalculated in Excel -- without needing to guess from blank grid
    # cells, which false-positived on legitimately-blank IF() formula
    # results (round-3 adversarial review finding P1; see
    # _dilmore_cache_still_stale's docstring for the full story). Written
    # from whichever caller triggers a real write (cmd_dilmore or
    # cmd_deliver's auto-run), not duplicated in each caller.
    state = _load_state(workbook_path.parent)
    state["formula_cache_stale"] = True
    state["formula_cache_stale_mtime"] = workbook_path.stat().st_mtime
    _save_state(workbook_path.parent, state)
    wb_values.close()

    return {
        "ok": True,
        "count": len(row_idxs),
        "changed": True,
        "tab": tab_name,
        "message": f"Wrote {len(row_idxs)} size adjustment(s) to {tab_name}.",
        "subject_gba": subject_gba,
        "curve": curve,
        "summary": summary,
    }


def cmd_dilmore(args):
    """Compute Dilmore size adjustments and write them to the adjustment grid tab."""
    if not args:
        print("Usage: python axiom.py dilmore <file_no>")
        return

    file_no = args[0]
    adir    = _find_assignment(file_no)

    if not adir:
        print(f"  Assignment {file_no} not found.")
        return

    workbook_path = adir / "workbook.xlsx"
    if not workbook_path.exists():
        print(f"  workbook.xlsx not found in {adir.name}")
        return

    result = _run_dilmore_calc(workbook_path)

    if not result["ok"] or result["count"] == 0:
        print(f"  {result['message']}")
        return

    print(f"\n  Subject GBA: {result['subject_gba']:,.0f} SF")
    print(f"  Curve: {result['curve']}%")
    print()
    print(f"  {'Comp':<12} {'Comp GBA':>10}  {'Ratio':>8}  {'Factor':>8}  {'Adj %':>8}")
    print("  " + "─" * 52)
    for row in result["summary"]:
        label = f'Comp {row["comp"]}'
        print(
            f'  {label:<12} {row["comp_gba"]:>10,.0f}  {row["ratio"]:>8.3f}  '
            f'{row["factor"]:>8.4f}  {row["adj_pct"]:>+8.1f}%'
        )
    print(f"  {result['message']}")


# --- Extract command ---

def cmd_extract(args):
    """Scan a folder and extract comp data from all Word and Excel files.

    Usage:
      python axiom.py extract <file_no>
      python axiom.py extract /full/path/to/folder
    """
    if not args:
        print('Usage: python axiom.py extract <file_no | folder_path>')
        return

    target = args[0]
    folder = Path(target)
    if not folder.exists() or not folder.is_dir():
        adir = _find_assignment(target)
        if adir:
            folder = adir
        else:
            print(f'  Not found: {target}')
            return

    print(f'  Extracting from: {folder}')
    print('  ' + '-' * 60)

    from extractor import extract_from_assignment
    result = extract_from_assignment(str(folder))

    comps       = result.get('comps', [])
    lease_comps = result.get('lease_comps', [])
    narr_data   = result.get('narrative', {}).get('data', {})
    rent_roll_entries = result.get('rent_roll_entries', [])
    expense_records = result.get('expense_records', [])
    market_observations = result.get('market_observations', [])
    artifacts = result.get('artifacts', [])
    sources     = result.get('sources', [])
    warnings    = result.get('warnings', [])

    print(f'  Sources scanned: {len(sources)}')
    for s in sources:
        print(f'    - {Path(s).name}')

    print(f'  Sale comps: {len(comps)}')
    for i, c in enumerate(comps, 1):
        d     = c['data']
        addr  = d.get('address_street', '-')[:45]
        price = d.get('sale_price')
        psf   = d.get('price_per_sf')
        date  = d.get('sale_date', '-')
        src   = c.get('source', '')
        price_str = f'${price:,.0f}' if isinstance(price, (int, float)) else '-'
        psf_str   = f'${psf:.2f}/SF' if isinstance(psf, (int, float)) else '-'
        print(f'    {i}. {addr}')
        print(f'       {price_str}  {psf_str}  {date}  [{Path(src).name}]')

    print(f'  Lease comps: {len(lease_comps)}')
    for i, c in enumerate(lease_comps, 1):
        d    = c['data']
        addr = d.get('address_street', '-')[:45]
        rent = d.get('base_rent_psf')
        sf   = d.get('sf_leased')
        date = d.get('lease_date', '-')
        src  = c.get('source', '')
        rent_str = f'${rent:.2f}/SF' if isinstance(rent, (int, float)) else '-'
        sf_str   = f'{sf:,.0f} SF' if isinstance(sf, (int, float)) else '-'
        print(f'    {i}. {addr}')
        print(f'       {rent_str}  {sf_str}  {date}  [{Path(src).name}]')

    if narr_data:
        print(f'  Narrative fields: {len(narr_data)}')
        for k, v in list(narr_data.items())[:8]:
            print(f'    {k}: {str(v)[:60]}')

    print(f'  Rent-roll rows: {len(rent_roll_entries)}')
    print(f'  Operating-expense lines: {len(expense_records)}')
    print(f'  Market observations: {len(market_observations)}')
    print(f'  Source artifacts: {len(artifacts)}')

    if warnings:
        print(f'  Warnings ({len(warnings)}):')
        for w in warnings:
            print(f'    ! {w}')


def cmd_comp_ingest(args):
    """Extract and stage historical assignment folders for human review."""
    if not args:
        print("Usage: python axiom.py comp-ingest <historical-projects-root>")
        return False
    root = Path(args[0])
    if not root.is_dir():
        print(f"  Historical projects folder not found: {root}")
        return False
    from ingest import run_extraction

    run_extraction(root)
    return True


def cmd_ocr_cleanup(args):
    """Delete OCR page-review images no longer referenced by any staged or
    confirmed batch awaiting review/commit."""
    from pdf_financial_extractor import prune_ocr_pages

    deleted = prune_ocr_pages()
    print(f"\n  Deleted {deleted} orphaned OCR page image(s).\n")
    return True


def cmd_review_staged(args):
    """Review staged comparable records in the terminal."""
    from ingest import review_staged

    review_staged()
    return True


def cmd_comp_commit(args):
    """Commit confirmed comparable batches to the canonical database."""
    from ingest import commit_confirmed

    commit_confirmed()
    return True


def cmd_comp_search(args):
    """Search reviewed sale or lease comparables in the local database."""
    kind = "sale"
    filters = {}
    index = 0
    while index < len(args):
        option = args[index]
        if option == "--lease":
            kind = "lease"
            index += 1
            continue
        if option in {"--city", "--type", "--address"}:
            if index + 1 >= len(args):
                print(f"  Missing value for {option}")
                return False
            filters[{
                "--city": "city",
                "--type": "property_type",
                "--address": "address_contains",
            }[option]] = args[index + 1]
            index += 2
            continue
        print(f"  Unknown comp-search option: {option}")
        return False

    from db import search_lease_comps, search_sale_comps

    records = (
        search_lease_comps(**filters)
        if kind == "lease"
        else search_sale_comps(**filters)
    )
    print(f"\n  {len(records)} reviewed {kind} comp(s) found")
    for number, record in enumerate(records, 1):
        address = record.get("address_street") or "(no address)"
        city = record.get("address_city") or ""
        if kind == "sale":
            amount = record.get("sale_price")
            metric = f"${amount:,.0f}" if amount is not None else "-"
            date = record.get("sale_date") or "-"
        else:
            amount = record.get("base_rent_psf")
            metric = f"${amount:,.2f}/SF" if amount is not None else "-"
            date = record.get("lease_date") or "-"
        print(f"    {number}. {address}, {city}")
        print(f"       {metric}  {date}")
    return True


def cmd_financial_search(args):
    """Search reviewed rent-roll rows or operating expenses."""
    kind = "rent_roll"
    filters = {}
    index = 0
    while index < len(args):
        option = args[index]
        if option == "--expenses":
            kind = "expense"
            index += 1
            continue
        value_options = {
            "--tenant": "tenant_contains",
            "--as-of": "as_of_date",
            "--year": "period_year",
            "--category": "category_contains",
        }
        if option in value_options:
            if index + 1 >= len(args):
                print(f"  Missing value for {option}")
                return False
            value = args[index + 1]
            if option == "--year":
                try:
                    value = int(value)
                except ValueError:
                    print("  --year must be a whole year such as 2025")
                    return False
            filters[value_options[option]] = value
            index += 2
            continue
        print(f"  Unknown financial-search option: {option}")
        return False

    allowed = (
        {"period_year", "category_contains"}
        if kind == "expense"
        else {"tenant_contains", "as_of_date"}
    )
    incompatible = set(filters) - allowed
    if incompatible:
        print(f"  One or more filters do not apply to {kind} records.")
        return False

    from db import search_operating_expenses, search_rent_roll_entries

    records = (
        search_operating_expenses(**filters)
        if kind == "expense"
        else search_rent_roll_entries(**filters)
    )
    print(f"\n  {len(records)} reviewed {kind} record(s) found")
    for number, record in enumerate(records, 1):
        if kind == "expense":
            amount = record.get("amount")
            metric = f"${amount:,.0f}" if amount is not None else "-"
            print(
                f"    {number}. {record.get('category') or '(no category)'} "
                f"({record.get('period_year') or '-'}) — {metric}"
            )
        else:
            rent = record.get("monthly_rent")
            metric = f"${rent:,.0f}/mo" if rent is not None else "-"
            label = (
                record.get("tenant_name")
                or record.get("suite")
                or record.get("unit_id")
                or "(unidentified space)"
            )
            print(
                f"    {number}. {label} — "
                f"{record.get('sf_leased') or '-'} SF, {metric}"
            )
    return True


def cmd_observation_search(args):
    """Search reviewed historical market observations."""
    option_map = {
        "--category": "category",
        "--geography": "geography",
        "--type": "property_type",
        "--text": "text_contains",
        "--from": "effective_date_from",
        "--to": "effective_date_to",
    }
    filters = {}
    index = 0
    while index < len(args):
        option = args[index]
        if option not in option_map:
            print(f"  Unknown observation-search option: {option}")
            return False
        if index + 1 >= len(args):
            print(f"  Missing value for {option}")
            return False
        filters[option_map[option]] = args[index + 1]
        index += 2

    from db import search_market_observations

    records = search_market_observations(**filters)
    print(f"\n  {len(records)} reviewed market observation(s) found")
    for number, record in enumerate(records, 1):
        title = record.get("title") or "(untitled)"
        context = " · ".join(
            str(value)
            for value in (
                record.get("category"),
                record.get("geography"),
                record.get("effective_date"),
            )
            if value
        )
        excerpt = (record.get("observation_text") or "").replace("\n", " ")[:140]
        print(f"    {number}. {title}")
        print(f"       {context}")
        print(f"       {excerpt}")
    return True


def cmd_artifact_search(args):
    """Search reviewed maps, charts, photos, sketches, and exhibits."""
    option_map = {
        "--kind": "artifact_kind",
        "--title": "title_contains",
        "--geography": "geography",
        "--type": "property_type",
        "--sha256": "artifact_sha256",
    }
    filters = {}
    index = 0
    while index < len(args):
        option = args[index]
        if option not in option_map:
            print(f"  Unknown artifact-search option: {option}")
            return False
        if index + 1 >= len(args):
            print(f"  Missing value for {option}")
            return False
        filters[option_map[option]] = args[index + 1]
        index += 2

    from db import search_source_artifacts

    records = search_source_artifacts(**filters)
    print(f"\n  {len(records)} reviewed source artifact(s) found")
    for number, record in enumerate(records, 1):
        dimensions = (
            f"{record['width_px']}x{record['height_px']} px"
            if record.get("width_px") and record.get("height_px")
            else "-"
        )
        print(
            f"    {number}. {record.get('title') or '(untitled)'} "
            f"[{record.get('artifact_kind') or '-'}]"
        )
        print(
            f"       {dimensions}  "
            f"{record.get('artifact_filename') or '-'}  "
            f"{record.get('source_filename') or '-'}"
        )
    return True


def cmd_list(args):
    """List all assignments with their current stage."""
    dirs = sorted(
        (d for d in ASSIGNMENTS_DIR.iterdir() if d.is_dir()),
        key=lambda d: d.stat().st_mtime, reverse=True
    )
    if not dirs:
        print('  No assignments found.')
        return
    print(f'  {"File No":<14} {"Client":<30} {"Stage":<14} {"Last Modified"}')
    print('  ' + '-' * 72)
    for d in dirs:
        state = _load_state(d)
        parts = d.name.split('_', 1)
        fno    = parts[0] if parts else d.name
        client = parts[1].replace('_', ' ') if len(parts) > 1 else ''
        stage  = state.get('stage', 'new')
        mtime  = datetime.datetime.fromtimestamp(d.stat().st_mtime).strftime('%Y-%m-%d')
        print(f'  {fno:<14} {client:<30} {stage:<14} {mtime}')


def cmd_status(args):
    """Show detailed status for one assignment."""
    if not args:
        print('Usage: python axiom.py status <file_no>')
        return
    file_no = args[0]
    adir    = _find_assignment(file_no)
    if not adir:
        print(f'  Assignment {file_no} not found.')
        return
    state = _load_state(adir)
    print(f'  Assignment: {adir.name}')
    print(f'  Path:       {adir}')
    print(f'  Stage:      {state.get("stage", "new")}')
    for k, v in state.items():
        if k != 'stage':
            print(f'  {k}: {v}')
    files = list(adir.rglob('*'))
    print(f'  Files ({len(files)}):')
    for f in sorted(files):
        if f.is_file():
            print(f'    {f.relative_to(adir)}')


# ─── Delivery readiness check (shared by dashboard + future QC commands) ──────

def check_delivery_readiness(adir):
    """
    Dry-run the delivery template fill against an assignment's current
    variables (JSON + workbook outputs tab) WITHOUT writing any real
    output file. Returns:
      {'checked': True,  'missing': [...], 'blocks': [...]}   on success
      {'checked': False, 'reason': '...'}                     if it can't check
    Reuses fill_engine.fill_document exactly — same logic `deliver` uses,
    so what the dashboard shows always matches what `deliver` will do.
    """
    return validate_assignment(
        assignment_dir=adir,
        templates_dir=TEMPLATES_DIR,
        deliver_config=STAGES.get('deliver', {}),
        registry_path=FIELD_REGISTRY,
    )


def _dilmore_staleness_warnings(workbook_path):
    """Read-only check: does the adjustment-grid tab have a comp GBA
    entered with no corresponding Size Factor yet? That just means Dilmore
    hasn't run since that GBA was typed in -- routine, not an error, and it
    runs automatically on delivery. `validate` must stay non-mutating
    (see its docstring), so this only reads and never writes, unlike
    `_run_dilmore_calc`."""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(str(workbook_path), data_only=True)
    except Exception:
        return []

    tab_name = next(
        (name for name in _DILMORE_TAB_LAYOUTS if name in wb.sheetnames), None
    )
    if tab_name is None:
        return []

    layout = _DILMORE_TAB_LAYOUTS[tab_name]
    sa = wb[tab_name]

    # Header-driven column resolution, matching _run_dilmore_calc's write
    # path -- both used to hardcode column indices, which could silently
    # disagree about which column is which if a column was ever hand-
    # inserted (round-4 Fable adversarial review finding Q6). If the
    # expected headers aren't there, fail open (no warning) rather than
    # raise: this function only ever produces a soft, informational
    # warning, and _run_dilmore_calc's own "ok": False path is what
    # actually surfaces a real header-mismatch error to Derek.
    try:
        header_map = read_header_map(sa)
    except Exception:
        return []
    comp_gba_col = header_map.get(layout["comp_gba_header"])
    factor_col = header_map.get(layout["factor_header"])
    if comp_gba_col is None or factor_col is None:
        return []

    stale_comps = []
    # Same anchor-checked window as _run_dilmore_calc's write path (see its
    # comment) -- this staleness check must cover exactly the rows Dilmore
    # will actually write to, or a comp past row 16 would show no
    # staleness warning even though it will never get a Size Factor either
    # (Fable adversarial review finding N1), and a row whose anchor got
    # cleared/corrupted would wrongly still count as a live comp (finding
    # P3).
    if tab_name == "sca_adjustment_grid":
        scan_rows = []
        for candidate_row in range(7, GRID_MAX_DATA_ROW + 1):
            comp_label = sa.cell(row=candidate_row, column=1).value
            if not (isinstance(comp_label, str) and comp_label.startswith("Sale No.")):
                break
            scan_rows.append(candidate_row)
    else:
        scan_rows = list(range(7, 17))

    for row_idx in scan_rows:
        comp_gba = sa.cell(row=row_idx, column=comp_gba_col).value
        factor = sa.cell(row=row_idx, column=factor_col).value
        if comp_gba and not factor:
            stale_comps.append(str(row_idx - 6))

    if not stale_comps:
        return []
    return [
        f"{tab_name}: Comp {', '.join(stale_comps)} has a GBA entered but "
        f"no Size Factor yet -- runs automatically on delivery, or preview "
        f"it now with `python axiom.py dilmore <file_no>`."
    ]


def cmd_validate(args):
    """Check delivery readiness without writing assignment files or state."""
    if not args:
        print('Usage: python axiom.py validate <file_no>')
        return False

    file_no = args[0]
    adir = _find_assignment(file_no)
    if not adir:
        print(f'  Assignment {file_no} not found.')
        return False

    result = check_delivery_readiness(adir)
    workbook_path = adir / "workbook.xlsx"
    if workbook_path.exists():
        result['warnings'] = list(result.get('warnings', [])) + \
            _dilmore_staleness_warnings(workbook_path)
    print(f'\n  Delivery validation: {adir.name}')
    print('  ' + '-' * 60)
    if result.get("schema_version"):
        print(f'  Field registry: v{result["schema_version"]}')

    for error in result['errors']:
        print(f'  [ERROR] {error}')

    if result.get("stale_intake_fields"):
        stale_fields = result["stale_intake_fields"]
        print(f'\n  Stale exported Intake fields ({len(stale_fields)}):')
        for key in stale_fields:
            print(f'    - {key}')

    if result['missing']:
        print(f'\n  Missing required fields ({len(result["missing"])}):')
        for key in result['missing']:
            print(f'    - {key}')

    if result['unresolved_blocks']:
        print(f'\n  Unresolved document blocks ({len(result["unresolved_blocks"])}):')
        for key, reason in result['unresolved_blocks'].items():
            print(f'    - {key}: {reason}')

    if result['handled_blocks']:
        print(f'\n  Pipeline-handled blocks ({len(result["handled_blocks"])}):')
        for key in result['handled_blocks']:
            print(f'    - {key}')

    if result['warnings']:
        print(f'\n  Warnings ({len(result["warnings"])}):')
        for warning in result['warnings']:
            print(f'    [!] {warning}')

    if result['ready']:
        print('\n  [OK] READY TO DELIVER')
        return True

    print('\n  [X] NOT READY TO DELIVER')
    return False


def cmd_contract(args):
    """Audit workbook and template keys against the versioned field registry."""
    template_paths = []
    for stage in STAGES.values():
        for document in stage.get("documents", []):
            template_path = TEMPLATES_DIR / document.get("template", "")
            if template_path.exists():
                template_paths.append(template_path)

    result = audit_assignment_contract(
        registry_path=FIELD_REGISTRY,
        workbook_path=WORKBOOK_TPL,
        template_paths=template_paths,
        variables={},
    )

    print(f"\n  Field contract: v{result['schema_version'] or 'unavailable'}")
    try:
        registry = load_registry(FIELD_REGISTRY)
    except Exception:
        registry = None
    if registry:
        print(
            f"  Registered: {len(registry['fields'])} fields, "
            f"{len(registry['blocks'])} blocks"
        )
    for error in result["errors"]:
        print(f"  [ERROR] {error}")
    for warning in result["warnings"]:
        print(f"  [!] {warning}")

    if result["errors"]:
        print("\n  [X] CONTRACT DRIFT DETECTED")
        return False

    print("\n  [OK] Workbook and configured templates match the registry.")
    return True


# ─── Dashboard ────────────────────────────────────────────────────────────────

_STAGE_COLORS = {
    'new':      ('#EDEDED', '#555555'),
    'engaged':  ('#FCEFC7', '#8A6D00'),
    'delivered': ('#DCEFDD', '#1E6B2E'),
}

_DASHBOARD_CSS = """
  * { box-sizing: border-box; }
  body {
    font-family: Cambria, Georgia, serif;
    background: #F4F2F8;
    margin: 0;
    padding: 32px;
    color: #2B2530;
  }
  h1 {
    color: #3F2A66;
    font-size: 22px;
    margin: 0 0 4px 0;
  }
  .subtitle {
    color: #6B6270;
    font-size: 13px;
    margin: 0 0 28px 0;
  }
  .grid {
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(340px, 1fr));
    gap: 18px;
  }
  .card {
    background: #FFFFFF;
    border: 1px solid #D5CCEB;
    border-radius: 8px;
    padding: 18px 20px;
    box-shadow: 0 1px 3px rgba(63,42,102,0.08);
  }
  .card-header {
    display: flex;
    justify-content: space-between;
    align-items: flex-start;
    margin-bottom: 10px;
  }
  .file-no {
    font-size: 12px;
    color: #6B6270;
    letter-spacing: 0.02em;
  }
  .client {
    font-size: 17px;
    font-weight: bold;
    color: #3F2A66;
    margin: 2px 0 0 0;
  }
  .stage-badge {
    font-size: 11px;
    font-weight: bold;
    padding: 4px 10px;
    border-radius: 12px;
    text-transform: uppercase;
    letter-spacing: 0.03em;
    white-space: nowrap;
  }
  .property {
    font-size: 14px;
    margin: 6px 0;
    color: #2B2530;
  }
  .value-conclusion {
    font-size: 20px;
    font-weight: bold;
    color: #3F2A66;
    margin: 10px 0 2px 0;
  }
  .value-label {
    font-size: 11px;
    color: #6B6270;
    text-transform: uppercase;
    letter-spacing: 0.03em;
  }
  .meta-row {
    display: flex;
    justify-content: space-between;
    font-size: 12px;
    color: #6B6270;
    margin-top: 12px;
    padding-top: 10px;
    border-top: 1px solid #EDE8F5;
  }
  .files {
    margin-top: 10px;
    font-size: 12px;
  }
  .files a {
    color: #3F2A66;
    text-decoration: none;
    margin-right: 10px;
    display: inline-block;
    margin-bottom: 4px;
  }
  .files a:hover { text-decoration: underline; }
  .blockers {
    margin-top: 12px;
    padding: 10px 12px;
    border-radius: 6px;
    font-size: 12px;
  }
  .blockers.ready {
    background: #DCEFDD;
    color: #1E6B2E;
    font-weight: bold;
  }
  .blockers.notready {
    background: #FBEAEA;
    color: #9B2C2C;
  }
  .blockers.unknown {
    background: #F0F0F0;
    color: #6B6270;
    font-style: italic;
  }
  .blockers .count {
    font-weight: bold;
  }
  .blockers details summary {
    cursor: pointer;
    outline: none;
  }
  .blockers .key-list {
    margin: 6px 0 0 0;
    padding-left: 18px;
    font-size: 11px;
    line-height: 1.6;
    max-height: 160px;
    overflow-y: auto;
  }
  .empty {
    color: #6B6270;
    font-style: italic;
    padding: 40px;
    text-align: center;
  }
"""


def _dash_get(variables, key, default='—'):
    val = variables.get(key)
    if val in (None, '', 'N/A'):
        return default
    return val


def _blockers_html(adir, esc):
    check = check_delivery_readiness(adir)
    if not check['checked']:
        reasons = check.get('errors') or ['Unknown validation failure.']
        reason = '; '.join(str(item) for item in reasons)
        return f'<div class="blockers unknown">Delivery check unavailable — {esc(reason)}</div>'

    missing = check['missing']
    unresolved = check.get('unresolved_blocks', {})
    if not missing and not unresolved and not check.get('errors'):
        return '<div class="blockers ready">&#10003; Ready to deliver — no missing fields</div>'

    items = ''.join(f'<li>Missing field: {esc(k)}</li>' for k in missing)
    items += ''.join(
        f'<li>Unresolved block: {esc(k)}</li>'
        for k in unresolved
    )
    items += ''.join(
        f'<li>Validation error: {esc(str(error))}</li>'
        for error in check.get('errors', [])
    )
    count = len(missing) + len(unresolved) + len(check.get('errors', []))
    return (
        '<div class="blockers notready">'
        f'<details><summary><span class="count">{count} issue(s)</span> blocking delivery</summary>'
        f'<ul class="key-list">{items}</ul>'
        '</details>'
        '</div>'
    )


def _dash_card(adir, state):
    parts  = adir.name.split('_', 1)
    fno    = state.get('file_no') or (parts[0] if parts else adir.name)
    client = state.get('client') or (parts[1].replace('_', ' ') if len(parts) > 1 else '')
    stage  = state.get('stage', 'new')
    bg, fg = _STAGE_COLORS.get(stage, _STAGE_COLORS['new'])

    json_path = _find_json(adir)
    variables = {}
    if json_path and json_path.exists():
        try:
            with open(json_path, encoding='utf-8') as f:
                variables = json.load(f)
        except (json.JSONDecodeError, OSError):
            variables = {}

    address = _dash_get(variables, 'PROPERTY_ADDRESS', '')
    city    = _dash_get(variables, 'PROPERTY_CITY', '')
    subtype = _dash_get(variables, 'PROPERTY_SUBTYPE_FULL', '')
    prop_line = ', '.join(p for p in [address, city] if p and p != '—') or 'Property TBD'
    value_conclusion = _dash_get(variables, 'VALUE_CONCLUSION', None)
    approaches = _dash_get(variables, 'DEVELOPED_APPROACHES', '')

    esc = html_lib.escape

    out = []
    out.append('<div class="card">')
    out.append('  <div class="card-header">')
    out.append('    <div>')
    out.append(f'      <div class="file-no">{esc(str(fno))}</div>')
    out.append(f'      <div class="client">{esc(str(client))}</div>')
    out.append('    </div>')
    out.append(f'    <div class="stage-badge" style="background:{bg};color:{fg};">{esc(stage)}</div>')
    out.append('  </div>')
    out.append(f'  <div class="property">{esc(prop_line)}' +
               (f' &mdash; {esc(subtype)}' if subtype and subtype != '—' else '') + '</div>')

    if value_conclusion:
        out.append('  <div class="value-label">Value Conclusion</div>')
        out.append(f'  <div class="value-conclusion">{esc(str(value_conclusion))}</div>')
    if approaches and approaches != '—':
        out.append(f'  <div class="property" style="font-size:12px;color:#6B6270;">{esc(str(approaches))}</div>')

    # Actionable QC: what's blocking delivery right now, using the exact
    # same missing-key check that `axiom.py deliver` runs — no separate
    # logic to fall out of sync.
    out.append(_blockers_html(adir, esc))

    created   = state.get('created') or '—'
    engaged   = state.get('engaged') or '—'
    delivered = state.get('delivered') or '—'
    out.append('  <div class="meta-row">')
    out.append(f'    <span>Created<br><b>{esc(str(created))}</b></span>')
    out.append(f'    <span>Engaged<br><b>{esc(str(engaged))}</b></span>')
    out.append(f'    <span>Delivered<br><b>{esc(str(delivered))}</b></span>')
    out.append('  </div>')

    # File links (relative, so they work opened locally from axiom_platform/)
    rel_base = adir.relative_to(BASE_DIR)
    link_items = []
    wb_path = adir / 'workbook.xlsx'
    if wb_path.exists():
        link_items.append((str(rel_base / 'workbook.xlsx'), 'Workbook'))
    if json_path and json_path.exists():
        link_items.append((str(rel_base / json_path.name), 'Variables JSON'))
    outputs_dir = adir / 'outputs'
    if outputs_dir.exists():
        for f in sorted(outputs_dir.glob('*.docx')):
            link_items.append((str(rel_base / 'outputs' / f.name), f.stem.split('_', 1)[-1].replace('_', ' ')))

    if link_items:
        out.append('  <div class="files">')
        for href, label in link_items:
            href_escaped = esc(href.replace('\\', '/'))
            out.append(f'    <a href="{href_escaped}">{esc(label)}</a>')
        out.append('  </div>')

    out.append('</div>')
    return '\n'.join(out)


def cmd_dashboard(args):
    """Build dashboard.html — a visual overview of every assignment."""
    dirs = sorted(
        (d for d in ASSIGNMENTS_DIR.iterdir() if d.is_dir()),
        key=lambda d: d.stat().st_mtime, reverse=True
    )

    cards_html = ''
    if not dirs:
        cards_html = '<div class="empty">No assignments yet. Run: python axiom.py new &lt;file_no&gt; &lt;client&gt;</div>'
    else:
        cards = []
        for d in dirs:
            state = _load_state(d)
            cards.append(_dash_card(d, state))
        cards_html = '\n'.join(cards)

    generated = datetime.datetime.now().strftime('%B %d, %Y at %I:%M %p')

    page = []
    page.append('<!DOCTYPE html>')
    page.append('<html lang="en">')
    page.append('<head>')
    page.append('<meta charset="UTF-8">')
    page.append('<title>Axiom Commercial Appraisal — Dashboard</title>')
    page.append(f'<style>{_DASHBOARD_CSS}</style>')
    page.append('</head>')
    page.append('<body>')
    page.append('<h1>Axiom Commercial Appraisal — Assignment Dashboard</h1>')
    page.append(f'<div class="subtitle">Generated {generated} &middot; {len(dirs)} assignment(s) &middot; refresh with: python axiom.py dashboard</div>')
    page.append('<div class="grid">')
    page.append(cards_html)
    page.append('</div>')
    page.append('</body>')
    page.append('</html>')

    out_path = BASE_DIR / 'dashboard.html'
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(page))

    print(f'\n  Dashboard built: {len(dirs)} assignment(s)')
    print(f'  {out_path}')
    print(f'  Open it in a browser, or run this command again any time to refresh it.')


# Command dispatch
COMMANDS = {
    'new':       cmd_new,
    'engage':    cmd_engage,
    'deliver':   cmd_deliver,
    'validate':  cmd_validate,
    'contract':  cmd_contract,
    'dilmore':   cmd_dilmore,
    'extract':   cmd_extract,
    'comp-ingest': cmd_comp_ingest,
    'review-staged': cmd_review_staged,
    'comp-commit': cmd_comp_commit,
    'ocr-cleanup': cmd_ocr_cleanup,
    'comp-search': cmd_comp_search,
    'financial-search': cmd_financial_search,
    'observation-search': cmd_observation_search,
    'artifact-search': cmd_artifact_search,
    'list':      cmd_list,
    'status':    cmd_status,
    'dashboard': cmd_dashboard,
}


def main():
    if len(sys.argv) < 2 or sys.argv[1] in ('-h', '--help'):
        print('Axiom Commercial Appraisal Platform')
        print('Usage: python axiom.py <command> [args]')
        for name, fn in COMMANDS.items():
            doc = (fn.__doc__ or '').split('\n')[0].strip()
            print(f'  {name:<14} {doc}')
        return
    cmd  = sys.argv[1].lower()
    args = sys.argv[2:]
    if cmd not in COMMANDS:
        print(f'  Unknown command: {cmd}')
        sys.exit(1)
    result = COMMANDS[cmd](args)
    if result is False:
        sys.exit(1)



if __name__ == '__main__':
    main()
