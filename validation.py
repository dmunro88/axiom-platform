"""Non-mutating delivery validation for Axiom assignments."""

import contextlib
import datetime
import importlib.util
import io
import json
import os
import re
import tempfile
import zipfile
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
from docx.oxml.ns import qn
from lxml import etree

from adjustment_grid import GRIDS as ADJUSTMENT_GRID_SHEETS
from adjustment_grid import AdjustmentGridError, read_grid_rows
from comp_builder import load_comp_data
from field_registry import audit_assignment_contract, load_registry
from fill_engine import fill_document, load_variables
from media_blocks import MEDIA_BLOCKS, media_files_for_block, missing_media_reason
from structured_blocks import (
    OWNERSHIP_HISTORY_BLOCK,
    ownership_history_missing_fields,
)


NARRATIVE_BLOCKS = frozenset({
    "INSPECTION_NARRATIVE",
    "MARKET_AREA_OVERVIEW",
    "SCA_APPROACH_NARRATIVE",
    "SCA_ADJUSTMENT_NARRATIVE",
    "SCA_CONCLUSION_NARRATIVE",
    "CAP_RATE_NARRATIVE",
    "ENCUMBRANCES_NARRATIVE",
    "RECONCILIATION_NARRATIVE",
    "LAND_ADJUSTMENT_NARRATIVE",
})

EXCEL_ERROR_VALUES = frozenset({
    "#NULL!",
    "#DIV/0!",
    "#VALUE!",
    "#REF!",
    "#NAME?",
    "#NUM!",
    "#N/A",
    "#GETTING_DATA",
    "#SPILL!",
    "#CALC!",
})

PLACEHOLDER_PATTERN = re.compile(r"\[\[([A-Z0-9_]+)\]\]")


def _find_variables_json(assignment_dir):
    matches = sorted(Path(assignment_dir).glob("*_variables.json"))
    if len(matches) == 1:
        return matches[0], None
    if not matches:
        return None, "No *_variables.json file found."
    names = ", ".join(path.name for path in matches)
    return None, f"Multiple variables JSON files found: {names}"


def _formula_cache_findings(workbook_path, relevant_keys=None):
    """Return output-formula errors and warnings without modifying the workbook."""
    errors = []
    warnings = []
    restrict_to_relevant = relevant_keys is not None
    relevant_keys = set(relevant_keys or [])

    try:
        formula_wb = openpyxl.load_workbook(
            workbook_path, data_only=False, read_only=True
        )
        value_wb = openpyxl.load_workbook(
            workbook_path, data_only=True, read_only=True
        )
    except Exception as exc:
        return [f"Workbook could not be read: {exc}"], warnings

    try:
        if "outputs" not in formula_wb.sheetnames:
            return ["Workbook has no outputs sheet."], warnings

        formula_ws = formula_wb["outputs"]
        value_ws = value_wb["outputs"]
        if formula_ws.max_row is None:
            formula_ws.calculate_dimension(force=True)
        if value_ws.max_row is None:
            value_ws.calculate_dimension(force=True)
        uncached = []

        for row_number in range(2, formula_ws.max_row + 1):
            source_key = formula_ws.cell(row=row_number, column=2).value
            cached_key = value_ws.cell(row=row_number, column=2).value
            key = cached_key if cached_key is not None else source_key
            if key is None:
                continue
            key = str(key).strip()
            if key.startswith("="):
                key = f"outputs!B{row_number}"
            if not key or " " in key:
                continue
            if restrict_to_relevant and key not in relevant_keys:
                continue

            has_formula = False
            has_usable_cached_value = False
            key_errors = []
            for column in (3, 4):
                source_value = formula_ws.cell(
                    row=row_number, column=column
                ).value
                cached_value = value_ws.cell(
                    row=row_number, column=column
                ).value

                if isinstance(source_value, str) and source_value.startswith("="):
                    has_formula = True

                if (
                    isinstance(cached_value, str)
                    and cached_value.upper() in EXCEL_ERROR_VALUES
                ):
                    key_errors.append(
                        (
                            formula_ws.cell(
                                row=row_number,
                                column=column,
                            ).coordinate,
                            cached_value,
                        )
                    )
                elif cached_value not in (None, "", "None"):
                    has_usable_cached_value = True

            if key_errors:
                locations = ", ".join(
                    f"{coordinate} ({value})"
                    for coordinate, value in key_errors
                )
                errors.append(
                    f"{key} has Excel error cache(s) at outputs!{locations}."
                )
            elif has_formula and not has_usable_cached_value:
                uncached.append(key)

        if uncached:
            unique_keys = sorted(set(uncached))
            preview = ", ".join(unique_keys[:10])
            remainder = len(unique_keys) - 10
            suffix = f" (+{remainder} more)" if remainder > 0 else ""
            warnings.append(
                "Workbook contains formulas without cached Excel results: "
                f"{preview}{suffix}. Open and recalculate/save the workbook "
                "in Excel before relying on delivery output."
            )
    finally:
        formula_wb.close()
        value_wb.close()

    return errors, warnings


def _normalized_text(value):
    if value is None:
        return ""
    if isinstance(value, (datetime.date, datetime.datetime)):
        return value.isoformat()
    return " ".join(str(value).strip().split())


def _decimal_value(value, percentage=False):
    text = _normalized_text(value)
    if not text:
        return None
    has_percent = text.endswith("%")
    text = (
        text.removesuffix("%")
        .replace("$", "")
        .replace(",", "")
        .strip()
    )
    try:
        number = Decimal(text)
    except InvalidOperation:
        return None
    if percentage and has_percent:
        number /= Decimal("100")
    return number


def _date_value(value):
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    text = _normalized_text(value)
    for date_format in (
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%B %d, %Y",
        "%b %d, %Y",
    ):
        try:
            return datetime.datetime.strptime(text, date_format).date()
        except ValueError:
            continue
    return None


def _field_values_match(workbook_value, json_value, value_kind):
    workbook_text = _normalized_text(workbook_value)
    json_text = _normalized_text(json_value)
    if workbook_text == json_text:
        return True
    if not workbook_text or not json_text:
        return False

    if value_kind == "date_text":
        workbook_date = _date_value(workbook_value)
        json_date = _date_value(json_value)
        return (
            workbook_date is not None
            and json_date is not None
            and workbook_date == json_date
        )

    if value_kind in {"currency_text", "number_text"}:
        workbook_number = _decimal_value(workbook_value)
        json_number = _decimal_value(json_value)
        return (
            workbook_number is not None
            and json_number is not None
            and workbook_number == json_number
        )

    if value_kind == "percentage_text":
        workbook_number = _decimal_value(workbook_value, percentage=True)
        json_number = _decimal_value(json_value, percentage=True)
        return (
            workbook_number is not None
            and json_number is not None
            and workbook_number == json_number
        )

    return False


def intake_json_findings(workbook_path, json_path, registry_path):
    """Return canonical Intake keys that differ from the exported JSON."""
    registry = load_registry(registry_path)
    with open(json_path, encoding="utf-8") as json_file:
        exported = json.load(json_file)

    workbook = openpyxl.load_workbook(
        workbook_path,
        data_only=True,
        read_only=True,
    )
    try:
        if "Intake" not in workbook.sheetnames:
            return [], ["Workbook has no Intake sheet for JSON freshness checks."]
        sheet = workbook["Intake"]
        intake_values = {}
        for row in sheet.iter_rows(
            min_row=1,
            min_col=1,
            max_col=2,
            values_only=True,
        ):
            key = row[0]
            if isinstance(key, str) and re.fullmatch(
                r"[A-Z][A-Z0-9_]*",
                key.strip(),
            ):
                intake_values[key.strip()] = row[1]
    finally:
        workbook.close()

    stale = []
    for key, workbook_value in intake_values.items():
        definition = registry["fields"].get(key)
        if not definition or definition.get("source_of_truth") != "intake":
            continue
        json_value = exported.get(key)
        if not _field_values_match(
            workbook_value,
            json_value,
            definition.get("value_kind", "text"),
        ):
            stale.append(key)
    return sorted(stale), []


def _dilmore_cache_still_stale(assignment_dir, workbook_path):
    """Is the workbook's formula cache still in the state left by a real
    (changed=True) Dilmore write? See axiom.py's _run_dilmore_calc
    docstring: any such write's save() silently wipes every OTHER cached
    formula result workbook-wide (Net Adjustment, Indicated Value, the
    grid's own Time Adj %/Adjusted Price/Overall/Rating columns, the
    outputs tab, etc). cmd_deliver hard-stops on the SAME attempt that
    triggers the write, but a later attempt where Dilmore finds nothing
    new to write (changed=False) sailed through uncaught even though the
    cache from the earlier write was never actually recalculated in Excel
    (Fable adversarial review finding A1/N3).

    A first attempt at closing that gap inspected each grid row's known
    formula-output columns for a blank value -- but several of those
    columns are themselves conditional IF() formulas that legitimately
    cache as blank whenever an optional input is left empty (e.g. Time
    Adj % when Monthly Mkt Rate is blank, or Net Adjustment % on a land
    comp needing no adjustments at all). openpyxl reads a cached empty
    string back as the same None a wiped cache produces, so that
    row-by-row check couldn't tell a genuinely stale cache apart from an
    ordinary comp that simply doesn't need every adjustment -- a false
    positive that could permanently block a legitimate delivery, since
    pressing F9 does nothing to change an intentionally-blank IF() result
    (round-3 adversarial review finding P1).

    This replaces that per-row guess with the one signal that's actually
    trustworthy: cmd_deliver records exactly when a real write happened
    (state["formula_cache_stale"]) and the workbook's mtime at that exact
    moment (state["formula_cache_stale_mtime"]). If the workbook's current
    mtime still matches that recorded value, nothing has saved it since --
    the cache is still stale. If the mtime differs, the workbook has been
    saved again since (trusted to mean Derek let Excel recalculate first,
    per the on-screen instructions, the same trust-the-user assumption
    `_dilmore_staleness_warnings` already makes elsewhere) and the cache is
    no longer considered stale. Purely reads state and stats the file --
    writes nothing, keeping validate_assignment's non-mutating contract.
    """
    state_file = Path(assignment_dir) / ".axiom.json"
    if not state_file.exists():
        return False
    try:
        with open(state_file, encoding="utf-8") as f:
            state = json.load(f)
    except (json.JSONDecodeError, OSError):
        return False
    if not state.get("formula_cache_stale"):
        return False
    recorded_mtime = state.get("formula_cache_stale_mtime")
    if recorded_mtime is None:
        return False
    try:
        current_mtime = Path(workbook_path).stat().st_mtime
    except OSError:
        return False
    return current_mtime == recorded_mtime


def _classify_blocks(blocks, workbook_path, assignment_dir, variables):
    handled = []
    unresolved = {}

    anthropic_available = importlib.util.find_spec("anthropic") is not None
    anthropic_key_available = bool(os.environ.get("ANTHROPIC_API_KEY"))
    dilmore_cache_stale = _dilmore_cache_still_stale(assignment_dir, workbook_path)

    for block in sorted(blocks):
        if block == OWNERSHIP_HISTORY_BLOCK:
            missing_fields = ownership_history_missing_fields(variables)
            if missing_fields:
                unresolved[block] = (
                    "Ownership table requires values for: "
                    + ", ".join(missing_fields)
                    + "."
                )
            else:
                handled.append(block)
            continue

        if block in MEDIA_BLOCKS:
            if media_files_for_block(block, assignment_dir):
                handled.append(block)
            else:
                unresolved[block] = missing_media_reason(
                    block,
                    assignment_dir,
                )
            continue

        if block == "COMP_SHEETS_BLOCK":
            with contextlib.redirect_stdout(io.StringIO()):
                comps = load_comp_data(workbook_path)
            if comps:
                problems = []
                seen_numbers = set()
                for row_number, comp in enumerate(comps, start=1):
                    comp_number = comp.get("COMP_NO", "").strip()
                    normalized_number = comp_number.casefold()
                    if normalized_number in seen_numbers:
                        problems.append(
                            f"duplicate comparable number {comp_number!r}"
                        )
                    seen_numbers.add(normalized_number)
                    missing_core = [
                        label
                        for key, label in (
                            ("COMP_ADDRESS_LINE1", "address"),
                            ("COMP_SALE_PRICE", "sale price"),
                        )
                        if not comp.get(key, "").strip()
                    ]
                    if missing_core:
                        problems.append(
                            f"comparable row {row_number} is missing "
                            + " and ".join(missing_core)
                        )
                if problems:
                    unresolved[block] = (
                        "Comp data failed quality checks: "
                        + "; ".join(problems)
                        + "."
                    )
                else:
                    handled.append(block)
            else:
                unresolved[block] = (
                    "Comp-page handler is registered, but the workbook has no "
                    "usable comp_data rows."
                )
            continue

        if block in ADJUSTMENT_GRID_SHEETS:
            sheet_name = ADJUSTMENT_GRID_SHEETS[block]
            try:
                _, rows = read_grid_rows(workbook_path, sheet_name)
            except AdjustmentGridError as exc:
                unresolved[block] = str(exc)
                continue
            if not rows:
                unresolved[block] = (
                    f"Adjustment-grid handler is registered, but '{sheet_name}' "
                    "has no populated comp rows yet."
                )
                continue

            # A cached Excel formula-error token (#DIV/0!, #REF!, etc.) is
            # flagged by adjustment_grid._format_value as a
            # "[FORMULA ERROR -- ...]" string rather than rendered
            # verbatim -- but until now that flagged string still counted
            # as "handled" and shipped in the delivered report exactly
            # like a normal text cell (Fable adversarial review findings
            # A3/N2). Block delivery instead so the underlying formula
            # gets fixed first.
            error_comps = [
                row.get("Comp", "?")
                for row in rows
                if any(
                    isinstance(value, str) and value.startswith("[FORMULA ERROR")
                    for value in row.values()
                )
            ]
            if error_comps:
                unresolved[block] = (
                    f"'{sheet_name}' has a formula error in comp(s): "
                    f"{', '.join(error_comps)}. Fix the underlying Excel "
                    "formula and recalculate before delivering."
                )
                continue

            # See _dilmore_cache_still_stale's docstring: a real Dilmore
            # write wipes every cached formula result workbook-wide, and
            # this is the trustworthy (state + mtime based, not per-cell
            # guessing) way to tell whether that's still true right now.
            if dilmore_cache_stale:
                unresolved[block] = (
                    f"'{sheet_name}': the workbook's cached formula results "
                    "are still stale from Dilmore's last size-adjustment "
                    "write (Time Adj %/Adjusted Price/Net Adjustment/"
                    "Indicated Value/Overall/Rating and other formulas may "
                    "be blank or out of date). Open workbook.xlsx, let it "
                    "fully recalculate (e.g. press F9), save, then retry."
                )
                continue

            handled.append(block)
            continue

        if block in NARRATIVE_BLOCKS:
            if anthropic_available and anthropic_key_available:
                handled.append(block)
            elif not anthropic_available:
                unresolved[block] = (
                    "Narrative handler requires the anthropic Python package."
                )
            else:
                unresolved[block] = (
                    "Narrative handler requires ANTHROPIC_API_KEY."
                )
            continue

        unresolved[block] = "No pipeline handler is registered for this block."

    return handled, unresolved


def find_docx_placeholders(docx_path):
    """Return unresolved ``[[KEY]]`` placeholders found anywhere in a DOCX."""
    placeholders = set()
    with zipfile.ZipFile(docx_path) as package:
        for name in package.namelist():
            if not name.startswith("word/") or not name.endswith(".xml"):
                continue
            root = etree.fromstring(package.read(name))
            for paragraph in root.iter(qn("w:p")):
                text = "".join(
                    node.text or ""
                    for node in paragraph.iter(qn("w:t"))
                )
                placeholders.update(PLACEHOLDER_PATTERN.findall(text))
    return sorted(placeholders)


def validate_assignment(
    assignment_dir,
    templates_dir,
    deliver_config,
    registry_path=None,
):
    """
    Validate delivery readiness without writing assignment files or state.

    Warnings are advisory. Errors, missing ordinary placeholders, and unresolved
    block placeholders make ``ready`` false.
    """
    assignment_dir = Path(assignment_dir)
    templates_dir = Path(templates_dir)

    result = {
        "checked": False,
        "ready": False,
        "assignment": assignment_dir.name,
        "errors": [],
        "warnings": [],
        "missing": [],
        "blocks": [],
        "handled_blocks": [],
        "unresolved_blocks": {},
        "schema_version": None,
        "stale_intake_fields": [],
    }

    if not assignment_dir.is_dir():
        result["errors"].append(f"Assignment directory not found: {assignment_dir}")
        return result

    json_path, json_error = _find_variables_json(assignment_dir)
    if json_error:
        result["errors"].append(json_error)

    workbook_path = assignment_dir / "workbook.xlsx"
    if not workbook_path.exists():
        result["errors"].append("workbook.xlsx not found.")

    documents = deliver_config.get("documents", [])
    if not documents:
        result["errors"].append("No delivery document is configured.")
        return result

    template_name = documents[0].get("template")
    template_path = templates_dir / template_name if template_name else None
    if not template_name:
        result["errors"].append("Delivery template name is missing from config.")
    elif not template_path.exists():
        result["errors"].append(f"Delivery template not found: {template_name}")

    if result["errors"]:
        return result

    try:
        variables = load_variables(
            json_path=json_path,
            workbook_path=workbook_path,
        )
        optional_blank_keys = set()
        if registry_path:
            registry = load_registry(registry_path)
            optional_blank_keys = {
                key
                for key, definition in registry["fields"].items()
                if definition.get("required") is False
            }
        with tempfile.TemporaryDirectory() as temp_dir:
            scratch_path = Path(temp_dir) / "validation.docx"
            fill_result = fill_document(
                template_path,
                scratch_path,
                variables,
                optional_blank_keys=optional_blank_keys,
            )
    except Exception as exc:
        result["errors"].append(f"Template fill validation failed: {exc}")
        return result

    if registry_path:
        contract = audit_assignment_contract(
            registry_path=registry_path,
            workbook_path=workbook_path,
            template_paths=[template_path],
            variables=variables,
        )
        result["schema_version"] = contract["schema_version"]
        result["errors"].extend(contract["errors"])
        result["warnings"].extend(contract["warnings"])
        try:
            stale_fields, stale_warnings = intake_json_findings(
                workbook_path,
                json_path,
                registry_path,
            )
        except Exception as exc:
            result["errors"].append(
                f"Intake/JSON freshness check failed: {exc}"
            )
        else:
            result["stale_intake_fields"] = stale_fields
            result["warnings"].extend(stale_warnings)
            if stale_fields:
                result["errors"].append(
                    f"{len(stale_fields)} canonical Intake field(s) differ "
                    "from the exported JSON. Re-export JSON from the current "
                    "workbook."
                )

    formula_relevant_keys = set(fill_result["required_keys"])
    if registry_path:
        registry = load_registry(registry_path)
        formula_relevant_keys = {
            key
            for key in formula_relevant_keys
            if registry["fields"].get(key, {}).get("source_of_truth")
            == "workbook_output"
        }

    formula_errors, formula_warnings = _formula_cache_findings(
        workbook_path,
        relevant_keys=formula_relevant_keys,
    )
    result["errors"].extend(formula_errors)
    result["warnings"].extend(formula_warnings)

    result["checked"] = True
    result["missing"] = fill_result["missing"]
    result["blocks"] = fill_result["blocks"]

    handled, unresolved = _classify_blocks(
        result["blocks"], workbook_path, assignment_dir, variables
    )
    result["handled_blocks"] = handled
    result["unresolved_blocks"] = unresolved
    result["ready"] = not (
        result["errors"]
        or result["missing"]
        or result["unresolved_blocks"]
    )
    return result
